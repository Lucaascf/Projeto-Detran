# /home/lusca/py_excel/tkinter campssa/banco.py
import sqlite3
from frames.login_frame import UserManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Color, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox, Toplevel, Frame, Label
from tkcalendar import DateEntry
from datetime import datetime
import json
from funcoes_botoes import FuncoesBotoes
from planilhas import Planilhas
from typing import Optional, List, Dict, Any, Tuple
import logging
from database_connection import DatabaseConnection
import os
from auth.user_manager import UserManager


# Configuração de logging
logging.basicConfig(
    filename="database_operations.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


class DataBaseLogin:
    """
    SEÇÃO 1: INICIALIZAÇÃO E CONFIGURAÇÃO
    """

    # Inicializa o gerenciador do banco de dados
    def __init__(self, db_name: str = "login.db"):
        self.db_name = db_name
        self._current_user = None
        self.user_manager = UserManager(db_name)


    # Cria estrutura inicial do banco
    def create_db(self) -> None:
        """Cria o banco de dados de usuários se não existir"""
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user TEXT NOT NULL UNIQUE,
                        password TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """
                )
                conn.commit()
                logger.info("Tabela de usuários criada/verificada com sucesso")
        except sqlite3.Error as e:
            logger.error(f"Erro ao criar banco de dados de usuários: {e}")
            raise
    
    @property
    def current_user(self):
        return self._current_user

    @current_user.setter
    def current_user(self, value):
        self._current_user = value


    """
    SEÇÃO 2: OPERAÇÕES DE CRIAÇÃO E VALIDAÇÃO
    """

    # Cria novo usuário no sistema
    def create_user(self, username: str, password: str) -> bool:
        try:
            role = 'employee'
            permissions = UserManager.DEFAULT_PERMISSIONS['employee']
            
            if (self._current_user and 
                self.user_manager.current_user and 
                'manage_users' in self.user_manager.current_user.permissions):
                role = 'manager'
                permissions = UserManager.DEFAULT_PERMISSIONS['manager']

            return self.user_manager.create_user(username, password, role, permissions)
        except Exception as e:
            logging.error(f"Erro na criação do usuário: {e}")
            return False

    # Valida credenciais do usuário
    def validate_user(self, username: str, password: str) -> bool:
        try:
            authenticated_user = self.user_manager.authenticate(username, password)
            if authenticated_user:
                self._current_user = username
                return True
            return False
        except Exception as e:
            logging.error(f"Erro na validação do usuário: {e}")
            return False

    """
    SEÇÃO 3: OPERAÇÕES DE CONSULTA
    """

    # Busca usuário por nome
    def read_user(self, user):
        """Função para ser um usuário com base no user"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE user = ?", (user,))
        usuario = cursor.fetchone()
        conn.close()
        return usuario

    """
    SEÇÃO 4: OPERAÇÕES DE MODIFICAÇÃO
    """

    # Atualiza senha do usuário
    def update_user(self, user, new_password):
        """Função para atualizar a senha de um usuário"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET password = ? WHERE user =?", (new_password, user)
        )

    # Remove usuário do sistema
    def delete_user(self, user):
        """Função para deletar um usuário com base no user"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE user = ?", (user,))
        conn.comit()
        conn.close()

    @property
    def current_user(self):
        """Property para acessar usuário atual."""
        return self._current_user
    
    @current_user.setter
    def current_user(self, value):
        """Setter para usuário atual."""
        self._current_user = value

    def has_permission(self, permission: str) -> bool:
        if self._current_user and self.user_manager.current_user:
            return permission in self.user_manager.current_user.permissions
        return False

class DataBaseMarcacao:
    """
    SEÇÃO 1: INICIALIZAÇÃO E CONFIGURAÇÃO
    """

    # Inicializa o banco de dados de marcações
    def __init__(
        self,
        master: tk.Tk,
        planilhas: Planilhas,
        file_path: str,
        app: Any,
        db_name: str = "db_marcacao.db",
    ):
        """Inicializa o sistema de marcações."""
        # Configuração do logger
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)

        # Verifica se já existe um handler para evitar duplicação de logs
        if not self.logger.handlers:
            # Handler para arquivo
            fh = logging.FileHandler("database_operations.log")
            fh.setLevel(logging.INFO)

            # Handler para console
            ch = logging.StreamHandler()
            ch.setLevel(logging.INFO)

            # Formato do log
            formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
            fh.setFormatter(formatter)
            ch.setFormatter(formatter)

            # Adiciona os handlers ao logger
            self.logger.addHandler(fh)
            self.logger.addHandler(ch)

        self.db_name = db_name
        self.master = master
        self.create_db()
        self.funcoes_botoes = FuncoesBotoes(self.master, planilhas, file_path, app)

        # UI Components
        self.window: Optional[Toplevel] = None
        self.marcacoes_window: Optional[Toplevel] = None
        self.results_frame: Optional[Frame] = None
        self.date_entry: Optional[DateEntry] = None
        self.search_window: Optional[Toplevel] = None
        self.search_var: Optional[tk.StringVar] = None
        self.table_frame: Optional[Frame] = None

        # Form fields
        self.name_entry: Optional[tk.Entry] = None
        self.renach_entry: Optional[tk.Entry] = None
        self.phone_entry: Optional[tk.Entry] = None
        self.appointment_entry: Optional[DateEntry] = None
        self.observation_text: Optional[tk.Text] = None

        # Verifica marcações expiradas na inicialização
        self.check_expired_appointments()

    # Cria e atualiza estrutura do banco
    def create_db(self) -> None:
        """Cria e atualiza a estrutura do banco de dados"""
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()

                # Verifica se a tabela existe
                cursor.execute(
                    """
                    SELECT name FROM sqlite_master 
                    WHERE type='table' AND name='marcacoes'
                """
                )

                if cursor.fetchone() is None:
                    # Criação da tabela principal com nomes de colunas corrigidos
                    cursor.execute(
                        """
                        CREATE TABLE IF NOT EXISTS marcacoes (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            nome TEXT NOT NULL,
                            telefone TEXT,
                            renach TEXT NOT NULL UNIQUE,
                            data_agendamento TEXT NOT NULL,
                            observacao TEXT,
                            status_comparecimento TEXT DEFAULT 'pending',
                            historico_comparecimento TEXT DEFAULT '[]',
                            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        )
                    """
                    )

                    # Trigger para atualizar atualizado_em
                    cursor.execute(
                        """
                        CREATE TRIGGER IF NOT EXISTS update_marcacoes_timestamp 
                        AFTER UPDATE ON marcacoes
                        BEGIN
                            UPDATE marcacoes SET atualizado_em = CURRENT_TIMESTAMP 
                            WHERE id = NEW.id;
                        END;
                    """
                    )

                # Criação da tabela de alterações (nova)
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS marcacoes_changes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        renach TEXT NOT NULL,
                        tipo TEXT NOT NULL,
                        valor_anterior TEXT,
                        valor_novo TEXT,
                        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                        changes TEXT NOT NULL
                    )
                """
)

                # Índice para melhorar performance de busca (novo)
                cursor.execute(
                    """
                    CREATE INDEX IF NOT EXISTS idx_marcacoes_changes_renach 
                    ON marcacoes_changes(renach)
                """
                )

                conn.commit()
                logger.info("Banco de dados criado/atualizado com sucesso")

        except sqlite3.Error as e:
            logger.error(f"Erro ao criar/atualizar banco de dados: {e}")
            raise

    # Realiza migração do banco de dados
    def migrate_database(self) -> None:
        """Realiza a migração do banco de dados para a estrutura mais recente"""
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()

                # Primeiro fazemos backup da tabela atual
                cursor.execute(
                    """
                    CREATE TABLE IF NOT EXISTS marcacoes_backup AS 
                    SELECT * FROM marcacoes
                """
                )

                # Removemos a tabela antiga
                cursor.execute("DROP TABLE IF EXISTS marcacoes")

                # Criamos a nova tabela com a estrutura correta
                cursor.execute(
                    """
                    CREATE TABLE marcacoes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        nome TEXT NOT NULL,
                        renach TEXT NOT NULL UNIQUE,
                        telefone TEXT,
                        data_agendamento TEXT NOT NULL,
                        observacao TEXT,
                        status_comparecimento TEXT DEFAULT 'pending',
                        historico_comparecimento TEXT DEFAULT '[]',
                        criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """
                )

                # Tentamos migrar os dados antigos
                try:
                    cursor.execute(
                        """
                        INSERT INTO marcacoes (
                            nome, renach, telefone, data_agendamento, 
                            observacao, status_comparecimento, historico_comparecimento
                        )
                        SELECT 
                            nome, renach, telefone, data_agendamento,
                            observacao, status_comparecimento, historico_comparecimento
                        FROM marcacoes_backup
                    """
                    )
                except sqlite3.Error as e:
                    logger.error(f"Erro ao migrar dados antigos: {e}")

                # Criamos o trigger para atualização de timestamp
                cursor.execute(
                    """
                    CREATE TRIGGER IF NOT EXISTS update_marcacoes_timestamp 
                    AFTER UPDATE ON marcacoes
                    BEGIN
                        UPDATE marcacoes SET atualizado_em = CURRENT_TIMESTAMP 
                        WHERE id = NEW.id;
                    END;
                """
                )

                conn.commit()
                logger.info("Migração do banco de dados concluída com sucesso")

        except sqlite3.Error as e:
            logger.error(f"Erro durante a migração do banco: {e}")
            raise

    """
    SEÇÃO 2: MANIPULAÇÃO DE DADOS
    """

    # Verifica marcações expiradas
    def check_expired_appointments(self) -> None:
        """Verifica e atualiza o status de marcações expiradas"""
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()
                today = datetime.now().date().strftime("%Y-%m-%d")

                # Busca marcações pendentes com data anterior a hoje
                cursor.execute(
                    """
                    SELECT renach, data_agendamento, historico_comparecimento 
                    FROM marcacoes 
                    WHERE status_comparecimento = 'pending' 
                    AND data_agendamento < ?
                """,
                    (today,),
                )

                expired_appointments = cursor.fetchall()

                for renach, data_agendamento, historico_atual in expired_appointments:
                    try:
                        historico = (
                            json.loads(historico_atual) if historico_atual else []
                        )
                    except json.JSONDecodeError:
                        logger.warning(
                            f"Histórico inválido para RENACH {renach}, iniciando novo"
                        )
                        historico = []

                    # Adiciona entrada ao histórico
                    historico.append(
                        {
                            "data": data_agendamento,
                            "status": "missed",
                            "atualizado_em": datetime.now().strftime(
                                "%Y-%m-%d %H:%M:%S"
                            ),
                            "observacao": "Status atualizado automaticamente - Data expirada",
                        }
                    )

                    # Atualiza o status da marcação
                    cursor.execute(
                        """
                        UPDATE marcacoes 
                        SET status_comparecimento = 'missed',
                            historico_comparecimento = ?
                        WHERE renach = ?
                    """,
                        (json.dumps(historico), renach),
                    )

                    logger.info(
                        f"Marcação expirada atualizada para RENACH {renach}: {data_agendamento}"
                    )

                conn.commit()

        except sqlite3.Error as e:
            logger.error(f"Erro ao verificar marcações expiradas: {e}")
            raise

    # Formata número de telefone
    @staticmethod
    def format_phone(phone: str) -> str:
        """Formata número de telefone para (XX) XXXXX-XXXX ou (XX) XXXX-XXXX"""
        phone = "".join(filter(str.isdigit, phone))
        if len(phone) == 11:
            return f"({phone[:2]}) {phone[2:7]}-{phone[7:]}"
        elif len(phone) == 10:
            return f"({phone[:2]}) {phone[2:6]}-{phone[6:]}"
        return phone

    # Valida campos do formulário
    def validate_fields(self) -> bool:
        """Valida os campos obrigatórios do formulário"""
        if not all([self.name_entry, self.renach_entry]):
            logger.error("Campos de formulário não inicializados")
            messagebox.showerror("Erro", "Erro de inicialização do formulário")
            return False

        name = self.name_entry.get().strip()
        renach = self.renach_entry.get().strip()

        if not all([name, renach]):
            logger.warning("Tentativa de submissão com campos obrigatórios vazios")
            messagebox.showerror("Erro", "Nome e RENACH são obrigatórios!")
            return False

        if not renach.isdigit():
            logger.warning(f"RENACH inválido fornecido: {renach}")
            messagebox.showerror("Erro", "RENACH deve conter apenas números!")
            return False

        return True

    # Limpa campos do formulário
    def clear_fields(self) -> None:
        """Limpa todos os campos do formulário"""
        if all(
            [
                self.name_entry,
                self.renach_entry,
                self.phone_entry,
                self.observation_text,
            ]
        ):
            self.name_entry.delete(0, tk.END)
            self.renach_entry.delete(0, tk.END)
            self.phone_entry.delete(0, tk.END)
            self.observation_text.delete("1.0", tk.END)
            logger.info("Campos do formulário limpos")
        else:
            logger.warning("Tentativa de limpar campos não inicializados")

    """
    SEÇÃO 3: INTERFACE DE USUÁRIO
    """

    # Interface principal de adição
    def add_user(self):
        """Cria a interface para adicionar/atualizar paciente."""
        self.window = tk.Toplevel(self.master)
        self.window.title("Gerenciar Paciente")
        self.window.geometry("800x500")
        self.window.minsize(width=500, height=500)
        self.window.maxsize(width=500, height=500)

        # Configuração visual
        cor_fundo = self.master.cget("bg")
        cor_texto = "#ECF0F1"
        self.window.configure(bg=cor_fundo)

        # Frame principal
        main_frame = tk.Frame(self.window, bg=cor_fundo)
        main_frame.pack(expand=True, fill="both", padx=20, pady=10)

        # Título
        tk.Label(
            main_frame,
            text="Cadastro de Paciente",
            font=("Arial", 14, "bold"),
            bg=cor_fundo,
            fg=cor_texto,
        ).pack(pady=(0, 15))

        # Campos de entrada
        campos = [
            ("Nome:", "name_entry"),
            ("Renach:", "renach_entry"),
            ("Telefone:", "phone_entry"),
        ]

        for label_text, entry_name in campos:
            frame = tk.Frame(main_frame, bg=cor_fundo)
            frame.pack(fill="x", pady=5)

            tk.Label(
                frame, text=label_text, bg=cor_fundo, fg=cor_texto, width=10, anchor="w"
            ).pack(side="left")

            entry = tk.Entry(frame)
            entry.pack(side="left", expand=True, fill="x", padx=(0, 10))
            setattr(self, entry_name, entry)

        # Campo de data
        date_frame = tk.Frame(main_frame, bg=cor_fundo)
        date_frame.pack(fill="x", pady=5)

        tk.Label(
            date_frame, text="Data:", bg=cor_fundo, fg=cor_texto, width=10, anchor="w"
        ).pack(side="left")

        self.appointment_entry = DateEntry(
            date_frame,
            width=12,
            background="darkblue",
            foreground="white",
            borderwidth=2,
        )
        self.appointment_entry.pack(side="left")

        # Campo de observação
        tk.Label(
            main_frame, text="Observações:", bg=cor_fundo, fg=cor_texto, anchor="w"
        ).pack(fill="x", pady=(10, 5))

        self.observation_text = tk.Text(
            main_frame, height=4, wrap=tk.WORD, font=("Arial", 10)
        )
        self.observation_text.pack(fill="x", pady=(0, 10))

        # Botões
        button_frame = tk.Frame(main_frame, bg=cor_fundo)
        button_frame.pack(fill="x", pady=10)

        tk.Button(
            button_frame, text="Salvar", command=self.submit_patient, width=15
        ).pack(side="left", padx=5)

        tk.Button(
            button_frame, text="Limpar", command=self.clear_fields, width=15
        ).pack(side="left", padx=5)

        tk.Button(
            button_frame, text="Fechar", command=self.window.destroy, width=15
        ).pack(side="left", padx=5)

        self.funcoes_botoes.center(self.window)

    # Interface de histórico
    def view_marcacoes(self):
        """Cria a interface para visualização e gestão das marcações."""
        # Verifica marcações expiradas antes de mostrar a interface
        self.check_expired_appointments()

        # Configuração da janela principal
        self.marcacoes_window = tk.Toplevel(self.master)
        self.marcacoes_window.title("Gerenciador de Marcações")
        self.marcacoes_window.geometry("1700x700")
        cor_fundo = self.master.cget("bg")
        cor_texto = "#ECF0F1"
        self.marcacoes_window.configure(bg=cor_fundo)

        # Frame principal
        main_frame = tk.Frame(self.marcacoes_window, bg=cor_fundo)
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Título
        title_frame = tk.Frame(main_frame, bg=cor_fundo)
        title_frame.pack(fill="x", pady=(0, 20))

        tk.Label(
            title_frame,
            text="Gerenciador de Marcações",
            font=("Arial", 18, "bold"),
            bg=cor_fundo,
            fg=cor_texto,
        ).pack(side="left")

        # Frame de controles
        control_frame = tk.Frame(main_frame, bg=cor_fundo)
        control_frame.pack(fill="x", pady=(0, 20))

        # Frame para seleção de data
        date_frame = tk.Frame(control_frame, bg=cor_fundo)
        date_frame.pack(side="left")

        tk.Label(
            date_frame,
            text="Data:",
            bg=cor_fundo,
            fg=cor_texto,
            font=("Arial", 12, "bold"),
        ).pack(side="left", padx=(0, 10))

        self.date_entry = DateEntry(
            date_frame,
            width=12,
            background="darkblue",
            foreground="white",
            borderwidth=2,
            font=("Arial", 10),
        )
        self.date_entry.set_date(datetime.now().date())
        self.date_entry.pack(side="left")

        # Frame para busca
        search_frame = tk.Frame(control_frame, bg=cor_fundo)
        search_frame.pack(side="right")

        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.filter_marcacoes)

        tk.Label(
            search_frame,
            text="Buscar:",
            bg=cor_fundo,
            fg=cor_texto,
            font=("Arial", 12, "bold"),
        ).pack(side="left", padx=(0, 10))

        tk.Entry(
            search_frame, textvariable=self.search_var, width=30, font=("Arial", 10)
        ).pack(side="left")

        # Botão para ver histórico
        tk.Button(
            control_frame,
            text="Ver Histórico",
            command=self.view_detailed_history,
            bg="#3498db",
            fg="white",
            font=("Arial", 10),
        ).pack(side="right", padx=10)

        # Frame para resultados com scroll
        table_container = tk.Frame(main_frame, bg=cor_fundo)
        table_container.pack(fill="both", expand=True)

        # Canvas e scrollbars
        canvas = tk.Canvas(table_container, bg=cor_fundo)
        scrollbar_y = ttk.Scrollbar(
            table_container, orient="vertical", command=canvas.yview
        )
        scrollbar_x = ttk.Scrollbar(
            table_container, orient="horizontal", command=canvas.xview
        )

        self.results_frame = tk.Frame(canvas, bg=cor_fundo)
        canvas.create_window((0, 0), window=self.results_frame, anchor="nw")
        self.update_patient_list()

        # Configuração do canvas
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        # Empacotamento dos componentes de rolagem
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)

        # Atualiza a lista inicial
        self.update_patient_list()

        # Configuração de eventos
        self.date_entry.bind("<<DateEntrySelected>>", self.update_patient_list)
        canvas.bind_all(
            "<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"),
        )

        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        self.results_frame.bind("<Configure>", _on_frame_configure)

        # Centraliza a janela
        self.funcoes_botoes.center(self.marcacoes_window)

    # Abre janela de edição com formulário para alterar dados do paciente
    def edit_marcacao(self, patient):
        """Abre janela para edição de marcação."""
        edit_window = tk.Toplevel(self.marcacoes_window)
        edit_window.title("Editar Marcação")
        edit_window.geometry("400x500")
        edit_window.configure(bg=self.master.cget("bg"))

        # Configurações de cores
        cor_fundo = self.master.cget("bg")
        cor_texto = "#ECF0F1"

        # Frame principal
        main_frame = tk.Frame(edit_window, bg=cor_fundo)
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Título
        tk.Label(
            main_frame,
            text="Editar Dados da Marcação",
            font=("Arial", 14, "bold"),
            bg=cor_fundo,
            fg=cor_texto,
        ).pack(pady=(0, 20))

        # Frame para os campos
        fields_frame = tk.Frame(main_frame, bg=cor_fundo)
        fields_frame.pack(fill="x", pady=10)

        # Função para criar campos somente leitura
        def create_readonly_field(parent, label_text, default_value=""):
            frame = tk.Frame(parent, bg=cor_fundo)
            frame.pack(fill="x", pady=5)

            tk.Label(
                frame,
                text=label_text,
                font=("Arial", 10, "bold"),
                bg=cor_fundo,
                fg=cor_texto,
                width=12,
                anchor="w",
            ).pack(side="left")

            label = tk.Label(
                frame,
                text=default_value,
                font=("Arial", 10),
                bg="#2C3E50",
                fg=cor_texto,
                anchor="w",
                padx=5,
                pady=2,
            )
            label.pack(side="left", fill="x", expand=True)
            return label

        # Função para criar campos editáveis
        def create_editable_field(parent, label_text, default_value=""):
            frame = tk.Frame(parent, bg=cor_fundo)
            frame.pack(fill="x", pady=5)

            tk.Label(
                frame,
                text=label_text,
                font=("Arial", 10, "bold"),
                bg=cor_fundo,
                fg=cor_texto,
                width=12,
                anchor="w",
            ).pack(side="left")

            entry = tk.Entry(frame, font=("Arial", 10))
            entry.pack(side="left", fill="x", expand=True)
            entry.insert(0, default_value)
            return entry

        # Criar campos somente leitura
        nome_label = create_readonly_field(fields_frame, "Nome:", patient[0])
        renach_label = create_readonly_field(fields_frame, "RENACH:", patient[2])
        
        # Formatar data para exibição
        data_formatada = datetime.strptime(patient[5], "%Y-%m-%d").strftime("%d/%m/%Y") if patient[5] else ""
        data_label = create_readonly_field(fields_frame, "Data:", data_formatada)

        # Criar campos editáveis
        telefone_entry = create_editable_field(fields_frame, "Telefone:", patient[1])

        # Campo de observações
        tk.Label(
            fields_frame,
            text="Observações:",
            font=("Arial", 10, "bold"),
            bg=cor_fundo,
            fg=cor_texto,
        ).pack(anchor="w", pady=(10, 5))

        obs_text = tk.Text(fields_frame, height=4, font=("Arial", 10), wrap=tk.WORD)
        obs_text.pack(fill="x")
        obs_text.insert("1.0", patient[4] if patient[4] else "")

        def save_changes():
            """Salva as alterações no banco de dados."""
            telefone = telefone_entry.get().strip()
            observacoes = obs_text.get("1.0", tk.END).strip()

            try:
                conn = sqlite3.connect(self.db_name)
                cursor = conn.cursor()

                cursor.execute(
                    """
                    UPDATE marcacoes 
                    SET telefone = ?, observacao = ?
                    WHERE renach = ?
                """,
                    (telefone, observacoes, patient[2]),
                )

                conn.commit()
                messagebox.showinfo("Sucesso", "Dados atualizados com sucesso!")
                edit_window.destroy()
                self.update_patient_list()

            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao atualizar dados: {str(e)}")
            finally:
                conn.close()

        # Frame para botões
        button_frame = tk.Frame(main_frame, bg=cor_fundo)
        button_frame.pack(pady=20)

        tk.Button(
            button_frame,
            text="Salvar",
            command=save_changes,
            width=15,
            bg="#2ecc71",
            fg="white",
        ).pack(side="left", padx=5)

        tk.Button(
            button_frame,
            text="Cancelar",
            command=edit_window.destroy,
            width=15,
            bg="#e74c3c",
            fg="white",
        ).pack(side="left", padx=5)

        # Centraliza a janela
        self.funcoes_botoes.center(edit_window)

    # Abre janela com histórico completo e busca de pacientes
    def view_patient_history(self):
        """Interface para visualizar histórico completo de pacientes."""
        history_window = tk.Toplevel(self.master)
        history_window.title("Histórico de Pacientes")
        history_window.geometry("900x600")
        history_window.configure(bg=self.master.cget("bg"))

        cor_fundo = self.master.cget("bg")
        cor_texto = "#ECF0F1"

        # Frame principal
        main_frame = tk.Frame(history_window, bg=cor_fundo)
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Frame de busca
        search_frame = tk.Frame(main_frame, bg=cor_fundo)
        search_frame.pack(fill="x", pady=(0, 20))

        tk.Label(
            search_frame,
            text="Buscar por nome ou RENACH:",
            font=("Arial", 12, "bold"),
            bg=cor_fundo,
            fg=cor_texto,
        ).pack(side="left", padx=(0, 10))

        search_var = tk.StringVar()
        search_entry = tk.Entry(
            search_frame, textvariable=search_var, width=40, font=("Arial", 11)
        )
        search_entry.pack(side="left")

        # Frame para a tabela
        table_frame = tk.Frame(main_frame)
        table_frame.pack(fill="both", expand=True)

        # Treeview com scrollbar
        tree_scroll = ttk.Scrollbar(table_frame)
        tree_scroll.pack(side="right", fill="y")

        style = ttk.Style()
        style.configure(
            "Treeview",
            background=cor_fundo,
            fieldbackground=cor_fundo,
            foreground=cor_texto,
        )

        tree = ttk.Treeview(
            table_frame,
            columns=("Data", "Nome", "RENACH", "Status", "Observação"),
            show="headings",
            yscrollcommand=tree_scroll.set,
        )
        tree.pack(fill="both", expand=True)

        tree_scroll.config(command=tree.yview)

        # Configurar colunas
        tree.heading("Data", text="Data")
        tree.heading("Nome", text="Nome")
        tree.heading("RENACH", text="RENACH")
        tree.heading("Status", text="Status")
        tree.heading("Observação", text="Observação")

        for col in ("Data", "Nome", "RENACH", "Status", "Observação"):
            tree.column(col, width=150)

        def search_history(*args):
            search_term = search_var.get().strip().lower()
            if len(search_term) < 3:
                return

            # Limpa a tabela
            for item in tree.get_children():
                tree.delete(item)

            try:
                conn = sqlite3.connect(self.db_name)
                cursor = conn.cursor()

                # Busca pacientes que correspondam ao termo de busca
                cursor.execute(
                    """
                    SELECT name, renach, appointment_date, attendance_status, observation, attendance_history
                    FROM marcacoes
                    WHERE LOWER(name) LIKE ? OR LOWER(renach) LIKE ?
                    ORDER BY appointment_date DESC
                """,
                    (f"%{search_term}%", f"%{search_term}%"),
                )

                results = cursor.fetchall()

                for row in results:
                    nome, renach, data, status, obs, history = row

                    # Formata o status
                    status_text = {
                        "attended": "Compareceu",
                        "missed": "Não Compareceu",
                        "pending": "Pendente",
                    }.get(status, "Desconhecido")

                    # Insere a marcação atual
                    tree.insert(
                        "",
                        "end",
                        values=(
                            datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y"),
                            nome,
                            renach,
                            status_text,
                            obs or "",
                        ),
                    )

                    # Insere o histórico se existir
                    if history:
                        try:
                            hist_data = json.loads(history)
                            for entry in hist_data:
                                if isinstance(entry, dict):
                                    tree.insert(
                                        "",
                                        "end",
                                        values=(
                                            datetime.strptime(
                                                entry["date"], "%Y-%m-%d"
                                            ).strftime("%d/%m/%Y"),
                                            nome,
                                            renach,
                                            entry["status"],
                                            f"Histórico: {entry['updated_at']}",
                                        ),
                                    )
                        except json.JSONDecodeError:
                            pass

            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao buscar histórico: {str(e)}")
            finally:
                conn.close()

        search_var.trace("w", search_history)

        # Centraliza a janela
        self.funcoes_botoes.center(history_window)

    """
    SEÇÃO 4: OPERAÇÕES COM PACIENTES
    """

    # Processa envio de formulário
    def submit_patient(self) -> None:
        """Processa o envio do formulário de paciente"""
        if not self.validate_fields():
            return

        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()

                nome = self.name_entry.get().strip().upper()
                renach = self.renach_entry.get().strip()
                telefone = self.format_phone(self.phone_entry.get().strip())
                data_agendamento = self.appointment_entry.get_date().strftime(
                    "%Y-%m-%d"
                )
                observacao = self.observation_text.get("1.0", tk.END).strip()

                # Verifica existência do RENACH
                cursor.execute("SELECT id FROM marcacoes WHERE renach = ?", (renach,))

                if cursor.fetchone():
                    if messagebox.askyesno(
                        "Paciente Existente",
                        "Este RENACH já está cadastrado. Deseja atualizar a data da consulta?",
                    ):
                        cursor.execute(
                            """
                            UPDATE marcacoes 
                            SET data_agendamento = ?, 
                                observacao = ?
                            WHERE renach = ?
                        """,
                            (data_agendamento, observacao, renach),
                        )
                        logger.info(f"Marcação atualizada para RENACH: {renach}")
                        messagebox.showinfo("Sucesso", "Data da consulta atualizada!")
                else:
                    cursor.execute(
                        """
                        INSERT INTO marcacoes (
                            nome, renach, telefone, data_agendamento, observacao
                        ) VALUES (?, ?, ?, ?, ?)
                    """,
                        (nome, renach, telefone, data_agendamento, observacao),
                    )
                    logger.info(f"Nova marcação criada para RENACH: {renach}")
                    messagebox.showinfo("Sucesso", "Paciente cadastrado com sucesso!")

                conn.commit()
                self.clear_fields()

        except sqlite3.Error as e:
            logger.error(f"Erro ao submeter paciente: {e}")
            messagebox.showerror("Erro", "Erro ao processar operação. Verifique o log.")

    # Exibe o histórico detalhado de alterações de um paciente, ordenado cronologicamente
    def view_detailed_history(self, renach=None):
        """Exibe o histórico detalhado de alterações de um paciente, ordenado cronologicamente."""
        history_window = tk.Toplevel(self.master)
        history_window.title("Histórico do Paciente")
        history_window.geometry("1200x700")
        
        # Configurações visuais
        cor_fundo = self.master.cget("bg")
        cor_texto = "#ECF0F1"
        
        # Container principal
        container = tk.Frame(history_window, bg=cor_fundo, padx=20, pady=10)
        container.pack(fill="both", expand=True)
        
        # Cabeçalho com busca
        header_frame = tk.Frame(container, bg=cor_fundo)
        header_frame.pack(fill="x", pady=(0, 20))
        
        tk.Label(
            header_frame,
            text="Histórico de Alterações",
            font=("Arial", 16, "bold"),
            bg=cor_fundo,
            fg=cor_texto
        ).pack(side="left")
        
        # Campo de busca
        search_frame = tk.Frame(header_frame, bg=cor_fundo)
        search_frame.pack(side="right")
        
        tk.Label(
            search_frame,
            text="Buscar:",
            font=("Arial", 12),
            bg=cor_fundo,
            fg=cor_texto
        ).pack(side="left", padx=(0, 10))
        
        search_var = tk.StringVar(value=renach if renach else "")
        search_entry = tk.Entry(
            search_frame,
            textvariable=search_var,
            width=30,
            font=("Arial", 11)
        )
        search_entry.pack(side="left")
        
        # Tabela com scrollbar
        table_container = tk.Frame(container, bg=cor_fundo)
        table_container.pack(fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(table_container)
        scrollbar.pack(side="right", fill="y")
        
        # Configuração das colunas da tabela
        columns = [
            ("data_hora", "Data/Hora", 150),
            ("tipo_alteracao", "Tipo de Alteração", 150),
            ("renach", "RENACH", 100),
            ("nome", "Nome", 200),
            ("valor_anterior", "Valor Anterior", 150),
            ("valor_novo", "Novo Valor", 150),
            ("data_agendamento", "Agendamento", 100),
        ]
        
        tree = ttk.Treeview(
            table_container,
            columns=[col[0] for col in columns],
            show="headings",
            yscrollcommand=scrollbar.set
        )
        
        # Configuração visual das colunas
        for col_id, heading, width in columns:
            tree.heading(col_id, text=heading)
            tree.column(col_id, width=width)
        
        tree.pack(fill="both", expand=True)
        scrollbar.config(command=tree.yview)
        
        def carregar_historico(search_term=None):
            """Carrega o histórico do paciente ordenado cronologicamente."""
            tree.delete(*tree.get_children())
            
            if not search_term:
                return
                
            try:
                with DatabaseConnection(self.db_name) as conn:
                    cursor = conn.cursor()
                    
                    # Lista para armazenar todos os eventos
                    todos_eventos = []
                    
                    # Busca dados do paciente
                    cursor.execute("""
                        SELECT nome, renach, historico_comparecimento, 
                            data_agendamento, status_comparecimento, 
                            criado_em
                        FROM marcacoes 
                        WHERE nome LIKE ? OR renach LIKE ?
                    """, (f"%{search_term}%", f"%{search_term}%"))
                    
                    registros = cursor.fetchall()
                    
                    if not registros:
                        messagebox.showinfo("Aviso", "Nenhum registro encontrado.")
                        return
                    
                    for registro in registros:
                        nome, renach, historico_json, data_agend, status, criado_em = registro
                        data_agend_fmt = datetime.strptime(data_agend, "%Y-%m-%d").strftime("%d/%m/%Y")
                        
                        # Adiciona registro inicial
                        timestamp_inicial = datetime.strptime(criado_em, "%Y-%m-%d %H:%M:%S")
                        todos_eventos.append({
                            'timestamp': timestamp_inicial,
                            'data_hora': timestamp_inicial.strftime("%d/%m/%Y %H:%M"),
                            'tipo': "Cadastro Inicial",
                            'renach': renach,
                            'nome': nome,
                            'anterior': "-",
                            'novo': f"Agendamento: {data_agend_fmt}",
                            'agendamento': data_agend_fmt
                        })
                        
                        # Busca alterações
                        cursor.execute("""
                            SELECT tipo, valor_anterior, valor_novo, timestamp
                            FROM marcacoes_changes
                            WHERE renach = ?
                        """, (renach,))
                        
                        for alteracao in cursor.fetchall():
                            tipo, anterior, novo, timestamp = alteracao
                            timestamp_alteracao = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                            
                            # Traduz status se necessário
                            if tipo == "Alteração de Status":
                                status_dict = {
                                    'attended': 'Compareceu',
                                    'missed': 'Não Compareceu',
                                    'pending': 'Pendente'
                                }
                                anterior = status_dict.get(anterior, anterior)
                                novo = status_dict.get(novo, novo)
                            
                            todos_eventos.append({
                                'timestamp': timestamp_alteracao,
                                'data_hora': timestamp_alteracao.strftime("%d/%m/%Y %H:%M"),
                                'tipo': tipo,
                                'renach': renach,
                                'nome': nome,
                                'anterior': anterior,
                                'novo': novo,
                                'agendamento': data_agend_fmt
                            })
                        
                        # Processa histórico de comparecimento
                        if historico_json:
                            try:
                                historico = json.loads(historico_json)
                                for entrada in historico:
                                    if isinstance(entrada, dict):
                                        status = entrada.get('status', '')
                                        atualizado_em = entrada.get('atualizado_em', '')
                                        timestamp_hist = datetime.strptime(atualizado_em, "%Y-%m-%d %H:%M:%S")
                                        
                                        status_traduzido = {
                                            'attended': 'Compareceu',
                                            'missed': 'Não Compareceu',
                                            'pending': 'Pendente'
                                        }.get(status, status)
                                        
                                        todos_eventos.append({
                                            'timestamp': timestamp_hist,
                                            'data_hora': timestamp_hist.strftime("%d/%m/%Y %H:%M"),
                                            'tipo': "Atualização de Status",
                                            'renach': renach,
                                            'nome': nome,
                                            'anterior': "-",
                                            'novo': status_traduzido,
                                            'agendamento': data_agend_fmt
                                        })
                            except json.JSONDecodeError as e:
                                logger.error(f"Erro ao processar histórico: {e}")
                    
                    # Ordena todos os eventos por timestamp (mais recente primeiro)
                    todos_eventos.sort(key=lambda x: x['timestamp'], reverse=True)
                    
                    # Insere eventos ordenados na tabela
                    for evento in todos_eventos:
                        tree.insert("", "end", values=(
                            evento['data_hora'],
                            evento['tipo'],
                            evento['renach'],
                            evento['nome'],
                            evento['anterior'],
                            evento['novo'],
                            evento['agendamento']
                        ))
                                
            except sqlite3.Error as e:
                logger.error(f"Erro na consulta: {e}")
                messagebox.showerror("Erro", "Erro ao carregar histórico")
        
        # Configura busca com delay
        def delayed_search(*args):
            if hasattr(delayed_search, 'timer_id'):
                history_window.after_cancel(delayed_search.timer_id)
            delayed_search.timer_id = history_window.after(300, lambda: carregar_historico(search_var.get().strip()))
        
        search_var.trace('w', delayed_search)
        
        # Carrega histórico inicial se RENACH fornecido
        if renach:
            carregar_historico(renach)
        
        # Centraliza a janela
        self.funcoes_botoes.center(history_window)

    # Remove marcação
    def delete_marcacao(self, patient):
        """Remove uma marcação do banco de dados."""
        if messagebox.askyesno(
            "Confirmar Exclusão",
            f"Deseja realmente excluir a marcação de {patient[0]}?",
        ):
            try:
                conn = sqlite3.connect(self.db_name)
                cursor = conn.cursor()

                cursor.execute("DELETE FROM marcacoes WHERE renach = ?", (patient[2],))
                conn.commit()

                messagebox.showinfo("Sucesso", "Marcação excluída com sucesso!")
                self.update_patient_list()

            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao excluir marcação: {str(e)}")
            finally:
                conn.close()

    # Atualiza status
    def update_attendance_status(self, renach: str, status: str) -> None:
        """Atualiza o status de comparecimento do paciente"""
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()

                # Primeiro pega o status atual antes de atualizar
                cursor.execute(
                    "SELECT status_comparecimento, data_agendamento FROM marcacoes WHERE renach = ?",
                    (renach,)
                )
                result = cursor.fetchone()
                if not result:
                    logger.warning(f"RENACH não encontrado: {renach}")
                    return

                status_anterior, data_agendamento = result

                # Tradução dos status para registro
                status_traduzido = {
                    'attended': 'Compareceu',
                    'missed': 'Não Compareceu',
                    'pending': 'Pendente'
                }

                # Cria o registro de alteração
                registro_alteracao = {
                    'tipo': 'Alteração de Status',
                    'valor_anterior': status_traduzido.get(status_anterior, status_anterior),
                    'valor_novo': status_traduzido.get(status, status),
                    'data_evento': data_agendamento,
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }

                # Insere o registro de alteração
                cursor.execute(
                    """
                    INSERT INTO marcacoes_changes (renach, tipo, valor_anterior, valor_novo, timestamp, changes)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (renach, registro_alteracao['tipo'], registro_alteracao['valor_anterior'],
                    registro_alteracao['valor_novo'], registro_alteracao['timestamp'],
                    json.dumps(registro_alteracao))
                )

                # Atualiza o status e o histórico
                historico_atual = result[0] if len(result) > 2 else '[]'
                try:
                    historico = json.loads(historico_atual) if historico_atual else []
                except json.JSONDecodeError:
                    historico = []

                historico.append({
                    'data': data_agendamento,
                    'status': status,
                    'atualizado_em': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })

                cursor.execute(
                    """
                    UPDATE marcacoes 
                    SET status_comparecimento = ?, 
                        historico_comparecimento = ?
                    WHERE renach = ?
                    """,
                    (status, json.dumps(historico), renach)
                )

                conn.commit()
                logger.info(f"Status atualizado para RENACH {renach}: {status}")
                self.update_patient_list()

        except sqlite3.Error as e:
            logger.error(f"Erro ao atualizar status: {e}")
            messagebox.showerror("Erro", "Erro ao atualizar status. Verifique o log.")

    """
    SEÇÃO 5: BUSCA E FILTROS
    """

    # Filtra marcações
    def filter_marcacoes(self):
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()
                if self.search_var:
                    search_term = self.search_var.get().strip()
                    if search_term:
                        cursor.execute(
                            "SELECT * FROM marcacoes WHERE nome LIKE ? OR telefone LIKE ? OR renach LIKE ?",
                            (f"%{search_term}%", f"%{search_term}%", f"%{search_term}%")
                        )
                    else:
                        cursor.execute("SELECT * FROM marcacoes")
                else:
                    cursor.execute("SELECT * FROM marcacoes")

                return cursor.fetchall()
        except sqlite3.Error as e:
            logger.error(f"Erro ao filtrar marcações: {e}")
            return []

    # Busca pacientes
    def get_patients_by_name_or_renach(
        self, search_term: str, selected_date: Optional[str] = None
    ) -> List[Tuple]:
        """Busca pacientes por nome ou RENACH"""
        try:
            with DatabaseConnection(self.db_name) as conn:
                cursor = conn.cursor()
                search_term = search_term.lower() if search_term else ""

                query = """
                    SELECT nome, telefone, renach, 
                        COALESCE(status_comparecimento, 'pending') as status_comparecimento, 
                        observacao, data_agendamento
                    FROM marcacoes 
                    WHERE (LOWER(nome) LIKE ? OR LOWER(renach) LIKE ?)
                """
                params = [f"%{search_term}%", f"%{search_term}%"]

                if selected_date and not search_term:
                    query += " AND data_agendamento = ?"
                    params.append(selected_date)

                query += " ORDER BY nome"

                cursor.execute(query, params)
                return cursor.fetchall()

        except sqlite3.Error as e:
            self.logger.error(f"Erro ao buscar pacientes: {e}")
            return []

    # Atualiza lista de pacientes
    def update_patient_list(self, event=None):
        """Atualiza a lista de pacientes com status de comparecimento."""
        # Verifica marcações expiradas antes de atualizar a lista
        self.check_expired_appointments()

        search_term = self.search_var.get().strip()
        selected_date = self.date_entry.get_date().strftime("%Y-%m-%d")
        patients = self.get_patients_by_name_or_renach(search_term, selected_date)

        # Limpa a tabela anterior
        for widget in self.results_frame.winfo_children():
            widget.destroy()

        # Adiciona cabeçalho atualizado
        headers = [
            "Nome",
            "Telefone",
            "RENACH",
            "Status",
            "Observações",
            "Data",
            "Ações",
        ]
        for j, header in enumerate(headers):
            header_cell = tk.Label(
                self.results_frame,
                text=header,
                font=("Arial", 12, "bold"),
                bg=self.master.cget("bg"),
                fg="#ECF0F1",
                width=20,
                anchor="w",
            )
            header_cell.grid(row=0, column=j, padx=5, pady=2, sticky="ew")

        if not patients:
            tk.Label(
                self.results_frame,
                text="Nenhum paciente encontrado.",
                font=("Arial", 11),
                bg=self.master.cget("bg"),
                fg="#ECF0F1",
                pady=20,
            ).grid(row=1, column=0, columnspan=len(headers))
            return

        # Popula a tabela com os dados atualizados
        for i, patient in enumerate(patients, start=1):
            for j, info in enumerate(patient):
                # Ajusta o texto do status
                if j == 3:  # Coluna de status
                    status_text = {
                        "attended": "Compareceu",
                        "missed": "Não Compareceu",
                        "pending": "Pendente",
                    }.get(info, "Pendente")

                    # Cores diferentes para cada status
                    status_colors = {
                        "attended": "#2ecc71",
                        "missed": "#e74c3c",
                        "pending": "#f1c40f",
                    }
                    bg_color = status_colors.get(info, "#f1c40f")
                elif j == 5:  # Coluna de data
                    date_text = info if info else ""
                    if date_text:
                        date_text = datetime.strptime(date_text, "%Y-%m-%d").strftime(
                            "%d/%m/%Y"
                        )
                    bg_color = self.master.cget("bg")
                else:
                    date_text = info if info else ""
                    bg_color = self.master.cget("bg")

                cell = tk.Label(
                    self.results_frame,
                    text=date_text if j == 5 else status_text if j == 3 else info,
                    font=("Arial", 11),
                    bg=bg_color,
                    fg="#ECF0F1" if j != 3 else "#000000",
                    width=25 if j in [0, 4] else 15,
                    anchor="w",
                    wraplength=300 if j == 4 else None,
                )
                cell.grid(row=i, column=j, padx=5, pady=2, sticky="ew")

            # Frame para botões de ação
            action_frame = tk.Frame(self.results_frame, bg=self.master.cget("bg"))
            action_frame.grid(row=i, column=len(headers) - 1, padx=5, pady=2)

            # Botões de ação
            def create_edit_callback(p):
                return lambda: self.edit_marcacao(p)

            def create_delete_callback(p):
                return lambda: self.delete_marcacao(p)

            tk.Button(
                action_frame,
                text="Editar",
                command=create_edit_callback(patient),
                width=8,
            ).pack(side="left", padx=2)

            tk.Button(
                action_frame,
                text="Excluir",
                command=create_delete_callback(patient),
                width=8,
            ).pack(side="left", padx=2)

            # Botões de status
            tk.Button(
                action_frame,
                text="✓",
                command=lambda p=patient: self.update_attendance_status(
                    p[2], "attended"
                ),
                bg="#2ecc71",
                fg="white",
                width=3,
            ).pack(side="left", padx=2)

            tk.Button(
                action_frame,
                text="✗",
                command=lambda p=patient: self.update_attendance_status(p[2], "missed"),
                bg="#e74c3c",
                fg="white",
                width=3,
            ).pack(side="left", padx=2)

            tk.Button(
                action_frame,
                text="⟲",
                command=lambda p=patient: self.update_attendance_status(
                    p[2], "pending"
                ),
                bg="#f1c40f",
                fg="white",
                width=3,
            ).pack(side="left", padx=2)

        # Estatísticas
        stats_frame = tk.Frame(self.results_frame, bg=self.master.cget("bg"))
        stats_frame.grid(
            row=len(patients) + 1, column=0, columnspan=len(headers), pady=10
        )

        # Contagem de status
        attended_count = sum(1 for p in patients if p[3] == "attended")
        missed_count = sum(1 for p in patients if p[3] == "missed")
        pending_count = sum(1 for p in patients if p[3] == "pending")

        stats_text = f"Total: {len(patients)} | Compareceram: {attended_count} | Não Compareceram: {missed_count} | Pendentes: {pending_count}"
        tk.Label(
            stats_frame,
            text=stats_text,
            font=("Arial", 10),
            bg=self.master.cget("bg"),
            fg="#ECF0F1",
        ).pack()



class SistemaContas:
    """
    Sistema de gerenciamento de contas com interface gráfica e persistência em banco de dados e Excel.
    """

    def __init__(self, file_path: str, db_path: str = "contas.db", current_user=None):
        """
        Inicializa o sistema de contas.

        Args:
            file_path (str): Caminho para o arquivo Excel
            db_path (str): Caminho para o banco de dados SQLite
            current_user (str, optional): Usuário atual do sistema
        """
        # Configuração de logging
        logging.basicConfig(
            filename='sistema_contas.log',
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)

        self.file_path = file_path
        self.db_path = db_path
        self.current_user = current_user
        self.sheet_name = "Contas Fechamento"
        
        # Inicialização do sistema
        self.criar_sheet_se_nao_existir()
        self.criar_banco_dados()
        
        # Atributos da interface
        self.window = None
        self.date_entry = None
        self.info_entry = None
        self.valor_entry = None
        self.status_label = None

    def criar_banco_dados(self):
        """Cria as tabelas necessárias no banco de dados SQLite."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Tabela principal de contas
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS contas (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        data DATE NOT NULL,
                        descricao TEXT NOT NULL,
                        valor DECIMAL(10,2) NOT NULL,
                        categoria TEXT,
                        status TEXT DEFAULT 'PENDENTE',
                        usuario_criacao TEXT,
                        data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        data_atualizacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Tabela de histórico de alterações
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS historico_alteracoes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        conta_id INTEGER,
                        tipo_alteracao TEXT NOT NULL,
                        valor_anterior TEXT,
                        valor_novo TEXT,
                        usuario TEXT,
                        data_alteracao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (conta_id) REFERENCES contas(id)
                    )
                """)
                
                # Índices para melhorar performance
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS idx_contas_data 
                    ON contas(data)
                """)
                
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS idx_historico_conta_id 
                    ON historico_alteracoes(conta_id)
                """)
                
                self.logger.info("Banco de dados inicializado com sucesso")
                
        except sqlite3.Error as e:
            self.logger.error(f"Erro ao criar banco de dados: {e}")
            messagebox.showerror("Erro", f"Erro ao criar banco de dados: {str(e)}")

    def salvar_informacoes(self, data_escolhida: str, info: str, valor: str) -> bool:
        """
        Salva as informações no banco de dados e no Excel.
        """
        try:
            # Converte a data para o formato do banco
            data_formatada = datetime.strptime(data_escolhida, "%d/%m/%Y").date()
            
            # Limpa o valor e converte para float
            valor_limpo = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
            valor_float = float(valor_limpo)
            
            # Operações no banco de dados
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Insere na tabela de contas
                cursor.execute(
                    """
                    INSERT INTO contas (data, descricao, valor, usuario_criacao)
                    VALUES (?, ?, ?, ?)
                    """,
                    (data_formatada.strftime("%Y-%m-%d"), info, valor_float, self.current_user)
                )
                
                conta_id = cursor.lastrowid
                
                # Registra no histórico
                cursor.execute(
                    """
                    INSERT INTO historico_alteracoes 
                    (conta_id, tipo_alteracao, valor_novo, usuario)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        conta_id,
                        'CRIAÇÃO',
                        json.dumps({
                            "data": data_escolhida,
                            "info": info,
                            "valor": valor
                        }),
                        self.current_user
                    )
                )
                
                conn.commit()
                
            # Atualiza o Excel
            self.atualizar_excel(data_formatada, info, valor_float)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Erro ao salvar informações: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar informações: {str(e)}")
            return False
        
    def buscar_contas_por_periodo(self, data_inicial: str, data_final: str) -> list:
        """Busca contas por período no banco de dados."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Ajusta formato das datas para comparação
            data_inicial = datetime.strptime(data_inicial, "%Y-%m-%d").date()
            data_final = datetime.strptime(data_final, "%Y-%m-%d").date()
            
            cursor.execute("""
                SELECT data, descricao, valor, usuario_criacao, data_criacao
                FROM contas 
                WHERE date(data) BETWEEN date(?) AND date(?)
                ORDER BY data
            """, (data_inicial, data_final))
            
            return cursor.fetchall()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro na consulta: {str(e)}")
            return []
        finally:
            conn.close()

    def obter_historico_alteracoes(self, conta_id: int) -> list:
        """Obtém histórico de alterações de uma conta específica."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT tipo_alteracao, valor_anterior, valor_novo, usuario, data_alteracao
                FROM historico_alteracoes
                WHERE conta_id = ?
                ORDER BY data_alteracao DESC
            """, (conta_id,))
            
            return cursor.fetchall()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao buscar histórico: {str(e)}")
            return []
        finally:
            conn.close()

    def atualizar_conta(self, conta_id: int, nova_descricao: str, novo_valor: float) -> bool:
        """Atualiza informações de uma conta existente."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Busca dados anteriores
            cursor.execute("SELECT descricao, valor FROM contas WHERE id = ?", (conta_id,))
            desc_anterior, valor_anterior = cursor.fetchone()
            
            # Atualiza a conta
            cursor.execute("""
                UPDATE contas 
                SET descricao = ?, valor = ?, data_atualizacao = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (nova_descricao, novo_valor, conta_id))
            
            # Registra alteração no histórico
            cursor.execute("""
                INSERT INTO historico_alteracoes (
                    conta_id, tipo_alteracao, valor_anterior, valor_novo, usuario
                )
                VALUES (?, 'ATUALIZAÇÃO', ?, ?, ?)
            """, (
                conta_id,
                json.dumps({"descricao": desc_anterior, "valor": valor_anterior}),
                json.dumps({"descricao": nova_descricao, "valor": novo_valor}),
                self.current_user
            ))
            
            conn.commit()
            return True
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao atualizar conta: {str(e)}")
            return False
        finally:
            conn.close()

    def formatar_valor_digitado(self, event=None):
        """
        Formata o valor enquanto o usuário digita, mantendo apenas números e mantendo
        a posição do cursor para permitir digitação contínua.
        """
        if not hasattr(self, 'valor_entry'):
            return
            
        # Guarda a posição do cursor
        cursor_pos = self.valor_entry.index(tk.INSERT)
        
        # Obtém o valor atual e remove formatação
        valor = self.valor_entry.get().replace('R$', '').replace('.', '').replace(',', '').strip()
        
        # Se não houver valor, limpa o campo e retorna
        if not valor:
            self.valor_entry.delete(0, tk.END)
            self.valor_entry.insert(0, "R$ 0,00")
            self.valor_entry.icursor(6)
            return
        
        try:
            # Remove caracteres não numéricos
            valor = ''.join(filter(str.isdigit, valor))
            
            # Converte para float (centavos)
            valor_float = float(valor) / 100
            
            # Formata como moeda
            valor_formatado = f"R$ {valor_float:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')
            
            # Atualiza o valor mantendo o cursor na posição correta
            self.valor_entry.delete(0, tk.END)
            self.valor_entry.insert(0, valor_formatado)
            
            # Ajusta a posição do cursor
            if cursor_pos > 0:
                # Conta quantos pontos de milhar existem até a posição do cursor
                texto_ate_cursor = valor_formatado[:cursor_pos + 3]
                num_pontos = texto_ate_cursor.count('.')
                nova_pos = cursor_pos + num_pontos + 3  # +3 pelo "R$ "
                self.valor_entry.icursor(min(nova_pos, len(valor_formatado)))
            
        except ValueError:
            # Se houver erro na conversão, reinicia o campo
            self.valor_entry.delete(0, tk.END)
            self.valor_entry.insert(0, "R$ 0,00")
            self.valor_entry.icursor(6)

    def atualizar_excel(self, data: datetime.date, info: str, valor: float):
        """
        Atualiza a planilha Excel com as novas informações.
        
        Args:
            data (datetime.date): Data da conta
            info (str): Descrição da conta
            valor (float): Valor da conta
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb[self.sheet_name]
            
            # Encontra a última linha com dados
            ultima_linha = ws.max_row + 1
            for row in range(ws.max_row, 0, -1):
                if ws.cell(row=row, column=1).value is not None:
                    ultima_linha = row + 1
                    break
            
            # Adiciona os novos dados
            ws.cell(row=ultima_linha, column=1, value=data)
            ws.cell(row=ultima_linha, column=2, value=info)
            ws.cell(row=ultima_linha, column=3, value=valor)
            
            # Formata as células
            ws.cell(row=ultima_linha, column=1).number_format = 'DD/MM/YYYY'
            ws.cell(row=ultima_linha, column=3).number_format = 'R$ #,##0.00'
            
            wb.save(self.file_path)
            
        except Exception as e:
            self.logger.error(f"Erro ao atualizar Excel: {e}")
            raise

    # Mantém métodos existentes relacionados à interface...
    def criar_sheet_se_nao_existir(self):
        """Mantém compatibilidade com versão anterior."""
        if os.path.exists(self.file_path):
            wb = load_workbook(self.file_path)
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=self.sheet_name)
                ws.append(["DATA", "CONTAS", "VALOR"])
                wb.save(self.file_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["DATA", "CONTAS", "VALOR"])
            wb.save(self.file_path)

    def abrir_janela(self):
        """Abre a janela principal do sistema."""
        self.window = tk.Toplevel()
        self.window.title("Sistema de Gerenciamento de Contas")
        self.window.geometry("600x500")
        
        # Configurações da janela
        self.window.transient(self.window.master)
        self.window.grab_set()
        self.window.focus_set()
        
        # Cria interface
        self.criar_interface()
        
        # Centraliza a janela
        self.centralizar_janela(self.window)

    def criar_interface(self):
        """Cria a interface gráfica principal do sistema."""
        if not self.window:
            return
            
        # Configurações visuais
        style = ttk.Style()
        style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Status.TLabel', font=('Arial', 10))
        
        # Frame principal
        main_frame = ttk.Frame(self.window, padding="20")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configuração de grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)
        
        # Título
        ttk.Label(
            main_frame, 
            text="Sistema de Gerenciamento de Contas", 
            style='Header.TLabel'
        ).grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Campo de data
        ttk.Label(main_frame, text="Data:").grid(row=1, column=0, sticky="w", pady=5)
        self.date_entry = DateEntry(
            main_frame,
            width=20,
            date_pattern="dd/mm/yyyy",
            background="darkblue",
            foreground="white",
            borderwidth=2
        )
        self.date_entry.grid(row=1, column=1, sticky="we", padx=(5, 0), pady=5)
        
        # Campo de descrição
        ttk.Label(main_frame, text="Descrição:").grid(row=2, column=0, sticky="w", pady=5)
        self.info_entry = ttk.Entry(main_frame)
        self.info_entry.grid(row=2, column=1, sticky="we", padx=(5, 0), pady=5)
        
        # Campo de valor com formatação automática
        ttk.Label(main_frame, text="Valor (R$):").grid(row=3, column=0, sticky="w", pady=5)
        self.valor_entry = ttk.Entry(main_frame)
        self.valor_entry.grid(row=3, column=1, sticky="we", padx=(5, 0), pady=5)
        # Vincula a função de formatação aos eventos de digitação
        self.valor_entry.bind('<KeyRelease>', self.formatar_valor_digitado)
        
        # Frame de botões
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)
        
        # Botões
        ttk.Button(button_frame, text="Salvar", command=self.capturar_dados, width=20).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(button_frame, text="Limpar", command=self.limpar_campos, width=20).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(button_frame, text="Visualizar Contas", command=self.visualizar_contas, width=20).grid(row=1, column=0, padx=5, pady=5)
        ttk.Button(button_frame, text="Fechar", command=self.window.destroy, width=20).grid(row=1, column=1, padx=5, pady=5)
        
        # Status
        self.status_label = ttk.Label(main_frame, text="", style='Status.TLabel')
        self.status_label.grid(row=5, column=0, columnspan=2, sticky="we", pady=(10, 0))
        
        # Foco inicial
        self.info_entry.focus()

    def capturar_dados(self):
        """Captura e valida os dados do formulário antes de salvar."""
        if self.validar_campos():
            data = self.date_entry.get()
            info = self.info_entry.get()
            valor = self.valor_entry.get()
            
            if self.salvar_informacoes(data, info, valor):
                messagebox.showinfo("Sucesso", "Informações salvas com sucesso!")
                self.limpar_campos()

    def validar_campos(self):
        """
        Valida os campos do formulário, com tratamento especial para o campo de valor.
        """
        info = self.info_entry.get().strip()
        valor = self.valor_entry.get().strip()
        data = self.date_entry.get().strip()

        if not all([data, info, valor]):
            messagebox.showerror("Erro", "Todos os campos são obrigatórios!")
            return False

        try:
            # Remove a formatação monetária e converte para float
            valor_limpo = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
            float(valor_limpo)  # Tenta converter para confirmar que é um número válido
            return True
        except ValueError:
            messagebox.showerror("Erro", "O valor deve ser um número válido!")
            return False

    def limpar_campos(self):
        self.info_entry.delete(0, tk.END)
        self.valor_entry.delete(0, tk.END)

    def visualizar_contas(self):
        """Abre a janela de visualização de contas."""
        view_window = tk.Toplevel(self.window)
        view_window.title("Visualização de Contas")
        view_window.geometry("1000x700")
        
        # Frame para filtros
        filter_frame = ttk.Frame(view_window, padding="10")
        filter_frame.pack(fill="x", padx=20, pady=10)
        
        # Campos de data
        ttk.Label(filter_frame, text="Período:").pack(side="left", padx=(0, 5))
        data_inicial = DateEntry(filter_frame, width=12, background='darkblue', foreground='white')
        data_inicial.pack(side="left", padx=5)
        
        ttk.Label(filter_frame, text="até").pack(side="left", padx=5)
        data_final = DateEntry(filter_frame, width=12, background='darkblue', foreground='white')
        data_final.pack(side="left", padx=5)
        
        # Frame para a tabela
        table_frame = ttk.Frame(view_window, padding="10")
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Configuração da tabela - Removida a coluna categoria
        columns = ("Data", "Descrição", "Valor", "Usuário", "Última Atualização")
        tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        
        # Configuração das colunas
        column_widths = {
            "Data": 100,
            "Descrição": 400,
            "Valor": 150,
            "Usuário": 150,
            "Última Atualização": 180
        }
        
        for col, width in column_widths.items():
            tree.heading(col, text=col)
            tree.column(col, width=width)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Layout da tabela e scrollbars
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        
        # Configuração do grid
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Frame para totais
        totals_frame = ttk.Frame(view_window, padding="10")
        totals_frame.pack(fill="x", padx=20, pady=10)
        
        total_label = ttk.Label(totals_frame, text="Total: R$ 0,00", style='Header.TLabel')
        total_label.pack(side="right")
        
        def atualizar_tabela():
            """Atualiza os dados da tabela."""
            for item in tree.get_children():
                tree.delete(item)
                
            data_ini = data_inicial.get_date().strftime("%Y-%m-%d")
            data_fim = data_final.get_date().strftime("%Y-%m-%d")
            
            try:
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    # Query atualizada para corresponder à estrutura atual do banco
                    cursor.execute("""
                        SELECT 
                            data, descricao, valor,
                            usuario_criacao, data_atualizacao
                        FROM contas 
                        WHERE data BETWEEN ? AND ?
                        ORDER BY data DESC
                    """, (data_ini, data_fim))
                    
                    total_valor = 0
                    for row in cursor.fetchall():
                        data = datetime.strptime(row[0], "%Y-%m-%d").strftime("%d/%m/%Y")
                        valor = float(row[2])
                        valor_formatado = f"R$ {valor:,.2f}"
                        data_atualizacao = datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
                        
                        tree.insert("", "end", values=(
                            data,               # Data
                            row[1],             # Descrição
                            valor_formatado,    # Valor
                            row[3] or "-",      # Usuário
                            data_atualizacao    # Última Atualização
                        ))
                        
                        total_valor += valor
                    
                    total_label.configure(text=f"Total: R$ {total_valor:,.2f}")
                    
            except sqlite3.Error as e:
                self.logger.error(f"Erro ao buscar dados: {e}")
                messagebox.showerror("Erro", "Erro ao carregar dados")
        
        # Botão de atualização
        ttk.Button(
            filter_frame, 
            text="Buscar", 
            command=atualizar_tabela
        ).pack(side="left", padx=20)
        
        # Carrega dados iniciais
        atualizar_tabela()
        
        # Centraliza a janela
        self.centralizar_janela(view_window)

    def centralizar_janela(self, window):
        """Centraliza uma janela na tela."""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')

    def _create_or_get_style(self, wb, name, **properties):
        """Cria ou recupera um estilo existente."""
        if name in wb.named_styles:
            return wb.named_styles[name]
        
        style = NamedStyle(name=name)
        for prop, value in properties.items():
            setattr(style, prop, value)
        wb.add_named_style(style)
        return style

