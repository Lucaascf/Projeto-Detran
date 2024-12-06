# /home/lusca/py_excel/tkinter campssa/funcoes_botoes.py
from dataclasses import dataclass
from config import ConfigManager
import sys
from typing import Dict, List, Optional, Tuple
import tkinter as tk
from tkinter import (
    messagebox,
    filedialog,
    Frame,
    Label,
    Entry,
    Button,
    Toplevel,
    SINGLE,
    VERTICAL,
    HORIZONTAL,
    END,
    Listbox,
)
from tkinter import ttk
from functools import lru_cache
import logging
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import ssl
from datetime import datetime
from tkcalendar import DateEntry
import sqlite3
from frames.ntfs_frame import EmitirNota
from database_connection import DatabaseConnection


# Configurando logs
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

"""Rola a página até um elemento e clica nele"""


def role_e_click(driver, xpath):
    """Rola a página até um elemento e clica nele.

    Args:
        driver: Instância do WebDriver do Selenium.
        xpath: XPath do elemento a ser clicado.
    """
    element = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, xpath))
    )
    # Rola para o elemento
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    element.click()


"""
Responsável por gerenciar todas as funcionalidades relacionadas à interface gráfica e operações de 
uma aplicação de gerenciamento de pacientes para uma clínica médica. Ela lida com cadastro de 
pacientes, processamento de pagamentos, geração de relatórios, e integração com serviços externos 
como WhatsApp e email.
"""


class FuncoesBotoes:
    """
    SEÇÃO 1: INICIALIZAÇÃO E CONFIGURAÇÃO

    Esta seção contém métodos relacionados à inicialização da classe e configuração inicial de
    variáveis e componentes
    """

    def __init__(self, master, planilhas, file_path, app, current_user=None):
        """
        Inicializa a classe FuncoesBotoes com as configurações básicas.

        Args:
            master (tk.Tk): Janela principal da aplicação
            planilhas (PlanilhaManager): Gerenciador de planilhas
            file_path (str): Caminho do arquivo da planilha
            app (App): Instância principal da aplicação
            current_user (str, optional): Usuário atual do sistema
        """
        self.master = master
        self.planilhas = planilhas
        self.file_path = file_path
        self.app = app
        self.current_user = current_user
        self.login_frame = None
        self.criar_conta_frame = None
        self.logger = logging.getLogger(__name__)
        self.db_manager = DatabaseManager("db_marcacao.db", self.logger)
        self.emitir_nota = None
        self.driver = None
        self.primeira_conta = None
        self.segunda_conta = None

        # Variáveis para pagamento
        self._init_payment_vars()

        # Initialize entry attributes
        self.nome_entry = None
        self.renach_entry = None
        self.valor_entries = {}
        self.dinheiro_entry = None
        self.cartao_entry = None
        self.pix_entry = None

    # Código de inicialização de variáveis de pagamento...
    def _init_payment_vars(self):
        """Inicializa variáveis de pagamento"""
        self.forma_pagamento_var = tk.StringVar(value="")
        self.radio_var = tk.StringVar(value="")
        self.payment_vars = {
            "D": tk.IntVar(),
            "C": tk.IntVar(),
            "E": tk.IntVar(),
            "P": tk.IntVar(),
        }

    # Código de configuração de usuário...
    def set_current_user(self, user):
        """
        Define o usuário atual do sistema.

        Args:
            user (str): Nome do usuário a ser definido como atual
        """
        self.current_user = user

    # Código de configuração de frames...
    def configurar_frames(self, login_frame, criar_conta_frame):
        """
        Configura os frames de login e criação de conta.

        Args:
            login_frame (tk.Frame): Frame de login
            criar_conta_frame (tk.Frame): Frame de criação de conta
        """
        self.login_frame = login_frame
        self.criar_conta_frame = criar_conta_frame

    """
    SEÇÃO 2: INTERFACE GRÁFICA E WIDGETS

    Esta seção contém métodos relacionados à criação 
    e manipulação de elementos da interface gráfica
    """

    # Código de centralização...
    def center(self, window):
        """
        Centraliza uma janela na tela.

        Args:
            window (tk.Toplevel/tk.Tk): Janela a ser centralizada
        """
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f"{width}x{height}+{x}+{y}")
        window.deiconify()

    def mostrar_criar_conta(self):
        """
        Alterna a visibilidade do frame de login para o frame de criação de conta.
        """
        if self.login_frame and self.criar_conta_frame:
            self.login_frame.grid_forget()
            self.criar_conta_frame.grid(
                row=0, column=0, sticky="nsew", padx=20, pady=20
            )

    def voltar_para_login(self):
        """
        Alterna a visibilidade do frame de criação de conta para o frame de login.
        """
        if self.login_frame and self.criar_conta_frame:
            self.criar_conta_frame.grid_forget()
            self.login_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

    # Código de criação do frame de pagamento...
    def _create_payment_frame(self, parent, cor_fundo, cor_texto, cor_selecionado):
        """
        Cria o frame de pagamento com todas as opções disponíveis.

        Args:
            parent (tk.Widget): Widget pai onde o frame será criado
            cor_fundo (str): Cor de fundo do frame
            cor_texto (str): Cor do texto
            cor_selecionado (str): Cor de seleção dos elementos

        Returns:
            tk.LabelFrame: Frame de pagamento configurado
        """
        frame_pagamento = tk.LabelFrame(
            parent,
            text="Formas de Pagamento",
            bg=cor_fundo,
            fg=cor_texto,
            font=("Arial", 12, "bold"),
        )
        frame_pagamento.pack(padx=20, pady=10, fill="x")

        def on_payment_change():
            selected_count = sum(var.get() for var in self.payment_vars.values())
            for forma, entry in self.valor_entries.items():
                if selected_count > 1:
                    entry.config(state="normal")
                    if not entry.get():
                        entry.config(bg="#FFE5E5")
                else:
                    entry.delete(0, tk.END)
                    entry.config(state="disabled", bg="#F0F0F0")

        formas_pagamento = {"D": "Débito", "C": "Crédito", "E": "Espécie", "P": "PIX"}

        for codigo, nome in formas_pagamento.items():
            frame = tk.Frame(frame_pagamento, bg=cor_fundo)
            frame.pack(fill="x", padx=10, pady=2)

            cb = tk.Checkbutton(
                frame,
                text=nome,
                variable=self.payment_vars[codigo],
                bg=cor_fundo,
                fg=cor_texto,
                selectcolor=cor_selecionado,
                activebackground=cor_fundo,
                activeforeground=cor_texto,
                highlightthickness=0,
                command=on_payment_change,
            )
            cb.pack(side=tk.LEFT, padx=(0, 10))

            valor_entry = tk.Entry(frame, width=15, state="disabled")
            valor_entry.pack(side=tk.LEFT)
            self.valor_entries[codigo] = valor_entry

            # Atribuir as entradas aos atributos correspondentes
            if codigo == "D":
                self.dinheiro_entry = valor_entry
            elif codigo == "C":
                self.cartao_entry = valor_entry
            elif codigo == "P":
                self.pix_entry = valor_entry

            tk.Label(frame, text="R$", bg=cor_fundo, fg=cor_texto).pack(
                side=tk.LEFT, padx=(5, 0)
            )

        return frame_pagamento

    # Código do frame de radio buttons...
    def _create_radio_frame(self, cor_fundo, cor_texto, cor_selecionado):
        """
        Cria o frame com os radio buttons para seleção do tipo de atendimento.

        Args:
            cor_fundo (str): Cor de fundo do frame
            cor_texto (str): Cor do texto dos radio buttons
            cor_selecionado (str): Cor de seleção dos radio buttons

        Returns:
            tk.Frame: Frame contendo os radio buttons configurados
        """
        frame_radios = tk.Frame(self.adicionar_window, bg=cor_fundo)
        frame_radios.pack(pady=5)

        tipos = [("Médico", "medico"), ("Psicólogo", "psicologo"), ("Ambos", "ambos")]
        for tipo, valor in tipos:
            tk.Radiobutton(
                frame_radios,
                text=tipo,
                variable=self.radio_var,
                value=valor,
                bg=cor_fundo,
                fg=cor_texto,
                selectcolor=cor_selecionado,
                activebackground=cor_fundo,
                activeforeground=cor_texto,
                highlightthickness=0,
                font=("Arial", 12),
            ).pack(side=tk.LEFT, padx=2)

    # Código do frame de botões...
    def _create_button_frame(self, cor_fundo):
        """
        Cria o frame com os botões de ação (Adicionar e Voltar).

        Args:
            cor_fundo (str): Cor de fundo do frame

        Returns:
            tk.Frame: Frame contendo os botões configurados
        """
        frame_botoes = tk.Frame(self.adicionar_window, bg=cor_fundo)
        frame_botoes.pack(pady=20)

        tk.Button(
            frame_botoes,
            text="Adicionar",
            command=self.salvar_informacao,
            width=15,
            highlightthickness=0,
            activebackground="#2C3E50",
            activeforeground="#ECF0F1",
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            frame_botoes,
            text="Voltar",
            command=self.adicionar_window.destroy,
            width=15,
            activebackground="#2C3E50",
            activeforeground="#ECF0F1",
        ).pack(side=tk.LEFT, padx=5)

    # Código de criação de entry...
    def criar_entry(self, frame_nome, var_name, parent):
        """
        Cria um frame com label e entry para entradas de texto.

        Args:
            frame_nome (str): Nome do label associado ao entry
            var_name (str): Nome da variável que armazenará o entry
            parent (tk.Widget): Widget pai onde o frame será criado

        Returns:
            tk.Entry: Widget de entrada criado
        """
        frame = tk.Frame(parent, bg=parent.cget("bg"))
        frame.pack(pady=2)

        tk.Label(
            frame,
            text=frame_nome,
            bg=parent.cget("bg"),
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(side=tk.LEFT, anchor="w", padx=5)

        entry = tk.Entry(frame)
        entry.pack(side=tk.LEFT, padx=2)
        setattr(self, var_name, entry)
        return entry  # Return the entry widget

    # Configura a interface para adição de novo paciente
    def _setup_add_interface(self, cor_fundo, cor_texto, cor_selecionado):
        """
        Configura a interface completa para adição de novo paciente.

        Cria e organiza todos os widgets necessários para a interface de adição,
        incluindo campos de entrada, botões de rádio e frame de pagamento.

        Args:
            cor_fundo (str): Cor de fundo da interface
            cor_texto (str): Cor do texto
            cor_selecionado (str): Cor dos elementos selecionados
        """
        # Título
        tk.Label(
            self.adicionar_window,
            text="Preencha as informações:",
            bg=cor_fundo,
            fg=cor_texto,
            font=("Arial", 16, "bold"),
        ).pack(pady=(15, 5))

        # Frame para RadioButtons
        self._create_radio_frame(cor_fundo, cor_texto, cor_selecionado)

        # Label para mostrar valor da consulta
        self.valor_consulta_label = tk.Label(
            self.adicionar_window,
            text="Valor da consulta: R$ 0,00",
            bg=cor_fundo,
            fg=cor_texto,
            font=("Arial", 10, "bold"),
        )
        self.valor_consulta_label.pack(pady=5)

        # Função para atualizar o valor da consulta
        def atualizar_valor_consulta(*args):
            tipo_consulta = self.radio_var.get()
            try:
                valor = PaymentProcessor.calculate_service_value(tipo_consulta)
                valor_formatado = PaymentProcessor.format_currency(valor)
                self.valor_consulta_label.config(
                    text=f"Valor da consulta: {valor_formatado}"
                )
            except ValueError:
                self.valor_consulta_label.config(text="Valor da consulta: R$ 0,00")

        # Associar a função ao radio_var
        self.radio_var.trace("w", atualizar_valor_consulta)

        # Entradas para nome e Renach
        self.nome_entry = self.criar_entry("Nome:", "nome_entry", self.adicionar_window)
        self.renach_entry = self.criar_entry(
            "Renach:", "renach_entry", self.adicionar_window
        )

        # Checkbox para reexame
        self.reexame_var = tk.BooleanVar()
        reexame_frame = tk.Frame(self.adicionar_window, bg=cor_fundo)
        reexame_frame.pack(pady=2)

        tk.Checkbutton(
            reexame_frame,
            text="Reexame",
            variable=self.reexame_var,
            bg=cor_fundo,
            fg=cor_texto,
            selectcolor=cor_selecionado,
            activebackground=cor_fundo,
            activeforeground=cor_texto,
            font=("Arial", 12),
        ).pack(side=tk.LEFT)

        # Frame de pagamento
        self._create_payment_frame(
            self.adicionar_window, cor_fundo, cor_texto, cor_selecionado
        )

        def limpar_campos():
            self.nome_entry.delete(0, tk.END)
            self.renach_entry.delete(0, tk.END)
            self.reexame_var.set(False)

            # Limpar campos de forma de pagamento
            for entry in self.valor_entries.values():
                entry.delete(0, tk.END)
                entry.config(state="disabled", bg="#F0F0F0")

            # Desmarcar checkbuttons de forma de pagamento
            for var in self.payment_vars.values():
                var.set(0)

            # Limpar seleção dos RadioButtons
            self.radio_var.set("")

        def adicionar_paciente():
            if self.verificar_soma_pagamentos():
                if self.salvar_informacao():
                    if self.adicionar_window.winfo_exists():
                        limpar_campos()
                        self.adicionar_window.destroy()

        # Botões
        button_frame = tk.Frame(self.adicionar_window, bg=cor_fundo)
        button_frame.pack(pady=10)

        tk.Button(
            button_frame,
            text="Adicionar",
            command=adicionar_paciente,
            bg="#2980B9",
            fg="white",
            font=("Arial", 12),
            width=10,
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            button_frame,
            text="Cancelar",
            command=self.adicionar_window.destroy,
            bg="#95A5A6",
            fg="white",
            font=("Arial", 12),
            width=10,
        ).pack(side=tk.LEFT, padx=5)

        # Texto de ajuda
        tk.Label(
            self.adicionar_window,
            text="Obs.: Para múltiplas formas de pagamento, informe o valor de cada uma.",
            bg=cor_fundo,
            fg=cor_texto,
            font=("Arial", 9, "italic"),
        ).pack(pady=(0, 10))

        # Limpar campos após criar todos os widgets
        limpar_campos()

    """
    SEÇÃO 3: MANIPULAÇÃO DE DADOS
    Esta seção contém métodos relacionados ao processamento e validação de dados
    """

    # Código de obtenção do workbook...
    def get_active_workbook(self):
        """
        Obtém o workbook ativo atualizado.

        Returns:
            openpyxl.Workbook: Workbook ativo ou None se não houver
        """
        if self.planilhas:
            self.planilhas.reload_workbook()
            return self.planilhas.wb
        return None

    # Código de verificação de pagamentos...
    def verificar_soma_pagamentos(self) -> bool:
        """
        Verifica se a soma dos valores de pagamento está correta.

        Returns:
            bool: True se a soma estiver correta, False caso contrário
        """
        return self.validar_pagamentos() is not None

    # Valida os valores e formas de pagamento selecionados
    def validar_pagamentos(self) -> Optional[Tuple[List[str], float]]:
        """
        Valida os valores e formas de pagamento selecionados.

        Verifica se os valores informados correspondem ao valor total esperado
        e se as formas de pagamento foram corretamente selecionadas.

        Returns:
            Optional[Tuple[List[str], float]]: Tupla contendo lista de pagamentos formatados
            e valor total esperado se válido, None caso contrário

        Raises:
            ValueError: Se houver erro na validação dos valores
        """
        try:
            # Obtém o tipo de consulta selecionado
            tipo_consulta = self.radio_var.get()
            if not tipo_consulta:
                messagebox.showerror("Erro", "Selecione o tipo de atendimento")
                return None

            # Obtém valor esperado usando PaymentProcessor
            try:
                valor_esperado = PaymentProcessor.calculate_service_value(tipo_consulta)
            except ValueError as e:
                messagebox.showerror("Erro", str(e))
                return None

            # Verificar formas de pagamento selecionadas
            formas_selecionadas = {
                forma: var.get() for forma, var in self.payment_vars.items()
            }

            if not any(formas_selecionadas.values()):
                messagebox.showerror(
                    "Erro", "Selecione pelo menos uma forma de pagamento."
                )
                return None

            # Processar pagamentos
            pagamentos = {}
            num_formas_selecionadas = sum(formas_selecionadas.values())

            for codigo, selecionado in formas_selecionadas.items():
                if selecionado:
                    valor_str = self.valor_entries[codigo].get().strip()

                    if num_formas_selecionadas == 1:
                        # Se for única forma de pagamento, usa o valor total
                        pagamentos[codigo] = PaymentProcessor.format_currency(
                            valor_esperado
                        )
                    else:
                        # Múltiplas formas de pagamento
                        if not valor_str:
                            messagebox.showerror(
                                "Erro",
                                "Informe o valor para todas as formas de pagamento selecionadas",
                            )
                            return None
                        pagamentos[codigo] = valor_str

            # Valida o total dos pagamentos
            if num_formas_selecionadas > 1:
                try:
                    if not PaymentProcessor.validate_payment_total(
                        pagamentos, valor_esperado
                    ):
                        messagebox.showerror(
                            "Erro",
                            f"A soma dos valores deve ser igual ao valor da consulta (R$ {valor_esperado:.2f})",
                        )
                        return None
                except ValueError as e:
                    messagebox.showerror("Erro", str(e))
                    return None

            # Processa e formata os métodos de pagamento
            formatted_payments = PaymentProcessor.process_payment_methods(pagamentos)
            return formatted_payments, valor_esperado

        except Exception as e:
            self.logger.error(f"Erro na validação de pagamentos: {e}")
            messagebox.showerror("Erro", "Erro ao validar pagamentos")
            return None

    # Calcula valores totais por tipo de atendimento e método de pagamento
    def _calcular_valores_atendimentos(self) -> Dict[str, Dict[str, float]]:
        """
        Calcula valores totais por tipo de atendimento e método de pagamento.

        Processa os dados da planilha e calcula totais para atendimentos médicos
        e psicológicos, separando por forma de pagamento.

        Returns:
            Dict[str, Dict[str, float]]: Dicionário com totais calculados por tipo
            e método de pagamento
        """
        try:
            # Carrega os dados da planilha
            wb = self.get_active_workbook()
            if not wb:
                raise ValueError("Não foi possível carregar o workbook")

            ws = wb.active
            if not ws:
                raise ValueError("Não foi possível acessar a planilha ativa")

            totais = {
                "medico": {
                    "Débito": 0,
                    "Crédito": 0,
                    "Espécie": 0,
                    "PIX": 0,
                    "total": 0,
                    "pacientes": 0,
                },
                "psicologo": {
                    "Débito": 0,
                    "Crédito": 0,
                    "Espécie": 0,
                    "PIX": 0,
                    "total": 0,
                    "pacientes": 0,
                },
            }

            # Processa pagamentos médicos (colunas B-F)
            self._processar_pagamentos_por_tipo(
                ws, tipo="medico", col_nome="B", col_pagamento="F", totais=totais
            )

            # Processa pagamentos psicólogo (colunas H-L)
            self._processar_pagamentos_por_tipo(
                ws, tipo="psicologo", col_nome="H", col_pagamento="L", totais=totais
            )

            return totais

        except Exception as e:
            self.logger.error(f"Erro ao calcular valores: {str(e)}")
            return None

    # Processa os pagamentos de um tipo específico de atendimento e atualiza o dicionário de totais.
    def _processar_pagamentos_por_tipo(
        self,
        ws,
        tipo: str,
        col_nome: str,
        col_pagamento: str,
        totais: Dict[str, Dict[str, float]],
    ) -> None:
        """
        Processa os pagamentos de um tipo específico de atendimento e atualiza o dicionário de totais.

        Analisa cada linha da planilha para o tipo de atendimento especificado,
        processando os pagamentos e atualizando os totais correspondentes.

        Args:
            ws: Worksheet ativa da planilha
            tipo (str): Tipo de atendimento ('medico' ou 'psicologo')
            col_nome (str): Letra da coluna que contém os nomes
            col_pagamento (str): Letra da coluna que contém os pagamentos
            totais (Dict[str, Dict[str, float]]): Dicionário para acumular os totais

        Notes:
            - Ignora linhas com nomes vazios ou que contenham palavras-chave específicas
            - Suporta dois formatos de pagamento:
                1. Código único (ex: 'D' para Débito)
                2. Múltiplos valores (ex: 'D:100,65|C:48,00')
            - Atualiza contadores de pacientes e valores totais
            - Registra erros no logger se houver problemas no processamento
        """
        for row in range(3, ws.max_row + 1):
            nome = ws[f"{col_nome}{row}"].value
            if not nome or not isinstance(nome, str):
                continue

            nome = nome.strip()
            if any(x in nome.lower() for x in ["soma", "médico", "psicólogo", "total"]):
                continue

            pagamento = ws[f"{col_pagamento}{row}"].value
            if not pagamento:
                continue

            totais[tipo]["pacientes"] += 1
            valor_consulta = PaymentProcessor.calculate_service_value(tipo)
            totais[tipo]["total"] += valor_consulta

            # Processa o pagamento
            if pagamento in PaymentProcessor.PAYMENT_TYPES:
                metodo = PaymentProcessor.PAYMENT_TYPES[pagamento]
                totais[tipo][metodo] += valor_consulta
            else:
                # Aceita tanto | quanto / como separadores
                partes_pagamento = []
                if "|" in pagamento:
                    partes_pagamento = pagamento.split("|")
                elif "/" in pagamento:
                    partes_pagamento = pagamento.split("/")
                else:
                    partes_pagamento = [pagamento]

                for parte in partes_pagamento:
                    try:
                        if ":" not in parte:
                            continue

                        metodo, valor = parte.split(":", 1)
                        metodo = metodo.strip()
                        valor = valor.strip()

                        # Remove qualquer texto adicional após o valor
                        valor = valor.split("/")[0].split("|")[0].strip()

                        valor_float = PaymentProcessor.convert_currency_value(valor)
                        metodo_traduzido = PaymentProcessor.PAYMENT_TYPES[metodo]
                        totais[tipo][metodo_traduzido] += valor_float
                    except (ValueError, KeyError) as e:
                        self.logger.error(f"Erro ao processar pagamento '{parte}': {e}")
                        continue

    """
    SEÇÃO 4: OPERAÇÕES COM PACIENTES
    Esta seção contém métodos relacionados ao gerenciamento de pacientes
    """

    # Código de adição de informações...
    def adicionar_informacao(self):
        """
        Cria uma nova janela para adicionar informações de pacientes.
        Configura a interface para entrada de dados do paciente.
        """
        self.adicionar_window = tk.Toplevel(self.master)
        self.adicionar_window.title("Adicionar Paciente")
        self.adicionar_window.geometry("500x450")
        self.adicionar_window.minsize(500, 450)
        self.adicionar_window.maxsize(500, 450)

        cor_fundo = self.master.cget("bg")
        cor_texto = "#ECF0F1"
        cor_selecionado = "#2C3E50"

        self.adicionar_window.configure(bg=cor_fundo)
        self.center(self.adicionar_window)

        # Configuração da interface
        self._setup_add_interface(cor_fundo, cor_texto, cor_selecionado)

    # Código de salvamento de informações...

    def salvar_informacao(self):
        """
        Valida dados e coordena o salvamento no banco e na planilha.

        Returns:
            bool: True se o salvamento for bem-sucedido, False caso contrário
        """
        try:
            # Obter e validar dados dos campos
            nome = self.nome_entry.get().strip().upper()
            renach = self.renach_entry.get().strip()
            escolha = self.radio_var.get()

            # Validações
            if not self._validar_campos_obrigatorios(nome, renach, escolha):
                return False

            if not renach.isdigit():
                messagebox.showerror("Erro", "O RENACH deve ser um número inteiro.")
                return False

            # Validar pagamentos
            resultado_validacao = self.validar_pagamentos()
            if resultado_validacao is None:
                return False

            pagamentos, valor_esperado = resultado_validacao

            # Salvar dados
            try:
                if self.salvar_na_planilha(nome, renach, pagamentos, escolha):
                    self.formatar_planilha()  # Formata apos salvar na planilha
                    if self.db_manager.adicionar_paciente(
                        nome, renach, pagamentos, escolha
                    ):  # Salva no banco de dados
                        self.adicionar_window.destroy()
                        return True

                messagebox.showerror("Erro", "Não foi possível salvar as informações")
                return False

            except Exception as e:
                self.logger.error(f"Erro ao salvar dados: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao salvar informações: {str(e)}")
                return False

        except Exception as e:
            self.logger.error(f"Erro ao processar informações: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao processar informações: {str(e)}")
            return False

    # Valida se todos os campos obrigatórios do formulário foram preenchidos
    def _validar_campos_obrigatorios(
        self, nome: str, renach: str, escolha: str
    ) -> bool:
        """
        Valida se todos os campos obrigatórios do formulário foram preenchidos.

        Args:
            nome (str): Nome do paciente
            renach (str): Número do RENACH
            escolha (str): Tipo de atendimento selecionado

        Returns:
            bool: True se todos os campos estão preenchidos, False caso contrário
        """
        if not all([nome, renach, escolha]):
            messagebox.showerror(
                "Erro",
                "Por favor, preencha todos os campos obrigatórios (nome, RENACH e tipo de atendimento).",
            )
            return False
        return True

    # Código de exclusão...
    def excluir(self):
        """
        Remove informações de pacientes da planilha com base no RENACH fornecido.

        Cria uma interface gráfica para input do RENACH e gerencia o processo de
        exclusão do paciente tanto da seção médica quanto psicológica.

        Funcionalidades:
        - Busca o paciente pelo número RENACH
        - Remove dados mantendo a integridade da planilha
        - Move conteúdo subsequente para preencher espaços vazios
        - Atualiza a planilha após a exclusão

        Raises:
            Exception: Se houver erro durante o processo de exclusão
        """
        try:
            wb = self.get_active_workbook()
            ws = wb.active
            conn = sqlite3.connect("db_marcacao.db")
            cursor = conn.cursor()

            def realizar_exclusao():
                """
                Executa o processo de exclusão do paciente.

                Valida o RENACH informado, localiza o paciente na planilha
                e coordena o processo de remoção dos dados.

                Raises:
                    ValueError: Se o RENACH não for um número válido
                    Exception: Para outros erros durante a exclusão
                """
                try:
                    renach = int(renach_entry.get().strip())

                    def limpar_linha(row_num, start_col, end_col):
                        """
                        Limpa os valores de uma linha específica da planilha.

                        Args:
                            row_num (int): Número da linha a ser limpa
                            start_col (int): Coluna inicial
                            end_col (int): Coluna final

                        Notes:
                            Verifica células mescladas para evitar erros de formatação
                        """
                        for col in range(start_col, end_col + 1):
                            cell = ws.cell(row=row_num, column=col)
                            # Verifica se não é uma célula mesclada
                            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                cell.value = None

                    def mover_conteudo(start_row, start_col, end_col):
                        """
                        Move o conteúdo das células para cima após uma exclusão.

                        Realiza a movimentação de dados de forma segura, preservando
                        a formatação e estrutura da planilha.

                        Args:
                            start_row (int): Linha inicial para movimentação
                            start_col (int): Coluna inicial
                            end_col (int): Coluna final

                        Notes:
                            - Move dados de baixo para cima para evitar sobreposição
                            - Trata células mescladas adequadamente
                            - Limpa a última linha após a movimentação
                        """
                        max_row = ws.max_row
                        # Move de baixo para cima para evitar sobrescrever dados
                        for row in range(start_row, max_row):
                            for col in range(start_col, end_col + 1):
                                current_cell = ws.cell(row=row, column=col)
                                next_cell = ws.cell(row=row + 1, column=col)

                                # Só copia se a célula atual não for mesclada
                                if not isinstance(
                                    current_cell, openpyxl.cell.cell.MergedCell
                                ):
                                    if isinstance(
                                        next_cell, openpyxl.cell.cell.MergedCell
                                    ):
                                        current_cell.value = None
                                    else:
                                        current_cell.value = next_cell.value

                        # Limpa a última linha
                        limpar_linha(max_row, start_col, end_col)

                    def encontrar_paciente(col_renach):
                        """
                        Encontra a linha do paciente pelo número do RENACH.

                        Args:
                            col_renach (int): Número da coluna que contém o RENACH

                        Returns:
                            int or None: Número da linha se encontrado, None caso contrário

                        Notes:
                            Ignora células mescladas durante a busca
                        """
                        for row in range(3, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_renach)
                            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                if cell.value and str(cell.value).strip() == str(
                                    renach
                                ):
                                    return row
                        return None

                    # Excluir paciente do banco de dados
                    cursor.execute("DELETE FROM marcacoes WHERE renach = ?", (renach,))
                    conn.commit()  # confirma exclusão

                    # Procura nas seções de médico e psicólogo
                    linha_medico = encontrar_paciente(3)  # Coluna C
                    linha_psi = encontrar_paciente(9)  # Coluna I

                    alteracoes = False

                    if linha_medico:
                        mover_conteudo(linha_medico, 2, 6)  # Colunas B-F
                        alteracoes = True
                        messagebox.showinfo("Sucesso", "Removido da seção de médicos")

                    if linha_psi:
                        mover_conteudo(linha_psi, 8, 12)  # Colunas H-L
                        alteracoes = True
                        messagebox.showinfo(
                            "Sucesso", "Removido da seção de psicólogos"
                        )

                    if alteracoes:
                        wb.save(self.file_path)
                        excluir_window.destroy()
                    else:
                        messagebox.showwarning("Aviso", "RENACH não encontrado")

                    conn.close()  # Fecha a conexão com o banco de dados

                except ValueError:
                    messagebox.showerror(
                        "Erro", "Por favor, insira um RENACH válido (apenas números)"
                    )
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao excluir paciente: {str(e)}")

            # Interface da janela de exclusão
            excluir_window = tk.Toplevel(self.master)
            excluir_window.title("Excluir Paciente")
            excluir_window.geometry("400x150")
            excluir_window.resizable(False, False)
            excluir_window.configure(bg=self.master.cget("bg"))
            excluir_window.transient(self.master)
            excluir_window.grab_set()

            # Frame principal
            main_frame = tk.Frame(excluir_window, bg=self.master.cget("bg"))
            main_frame.pack(expand=True, fill="both", padx=20, pady=20)

            # Label
            tk.Label(
                main_frame,
                text="Informe o RENACH:",
                bg=self.master.cget("bg"),
                fg="#ECF0F1",
                font=("Arial", 14, "bold"),
            ).pack(pady=10)

            # Entry frame
            entry_frame = tk.Frame(main_frame, bg=self.master.cget("bg"))
            entry_frame.pack(fill="x", pady=5)

            renach_entry = tk.Entry(entry_frame, justify="center")
            renach_entry.pack(pady=5)
            renach_entry.focus()
            renach_entry.bind("<Return>", lambda e: realizar_exclusao())

            # Botões
            button_frame = tk.Frame(main_frame, bg=self.master.cget("bg"))
            button_frame.pack(pady=10)

            tk.Button(
                button_frame,
                text="Excluir",
                command=realizar_exclusao,
                bg="#ff4444",
                fg="white",
                font=("Arial", 10, "bold"),
                width=15,
            ).pack(side=tk.LEFT, padx=5)

            tk.Button(
                button_frame, text="Cancelar", command=excluir_window.destroy, width=15
            ).pack(side=tk.LEFT, padx=5)

            self.center(excluir_window)

            def on_closing():
                """
                Manipula o fechamento da janela de exclusão.

                Fecha o workbook e destroi a janela de forma segura,
                garantindo que os recursos sejam liberados corretamente.
                """
                wb.close()
                excluir_window.destroy()

            excluir_window.protocol("WM_DELETE_WINDOW", on_closing)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao iniciar exclusão: {str(e)}")

    # Código de exibição de informações...
    def exibir_informacao(self):
        """
        Exibe uma janela com informações detalhadas dos pacientes.

        Utiliza a classe PatientInfoDisplay para mostrar as informações
        em uma interface gráfica organizada.

        Returns:
            bool: True se a exibição for bem-sucedida, False caso contrário
        """
        try:
            # Cria uma instância do PatientInfoDisplay passando os parâmetros necessários
            display = PatientInfoDisplay(
                master=self.master,  # Janela principal
                planilhas=self.planilhas,  # Objeto que gerencia as planilhas
                logger=self.logger,  # Logger para registro de eventos
            )

            # Chama o método display para mostrar as informações
            display.display()

            return True

        except Exception as e:
            self.logger.error(f"Erro ao exibir informações: {str(e)}")
            messagebox.showerror(
                "Erro", f"Ocorreu um erro ao exibir as informações: {str(e)}"
            )
            return False

    """
    SEÇÃO 5: MANIPULAÇÃO DE PLANILHAS
    Esta seção contém métodos relacionados à manipulação de planilhas Excel
    """

    # Código de formatação...
    def formatar_planilha(self):
        """
        Formata a planilha preservando as informações necessárias.
        Aplica estilos, bordas e alinhamentos padronizados.

        Returns:
            bool: True se a formatação for bem-sucedida, False caso contrário
        """

        try:
            if not self.planilhas:
                return False

            self.planilhas.reload_workbook()
            ws = self.planilhas.get_active_sheet()

            if not ws:
                return False

            # Definir estilos
            thin_side = Side(style="thin")
            borda = Border(
                left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
            )
            font_bold = Font(name="Arial", bold=True, size=11, color="000000")
            font_regular = Font(name="Arial", size=11, color="000000")
            alignment_center = Alignment(horizontal="center", vertical="center")
            alignment_left = Alignment(horizontal="left", vertical="center")

            # Coletar dados existentes
            dados_medicos = []
            dados_psicologos = []
            max_row = ws.max_row + 1

            # Função auxiliar para verificar se é linha de total
            def is_total_row(valor):
                if not valor or not isinstance(valor, str):
                    return False
                return any(
                    palavra in valor.lower()
                    for palavra in ["soma", "médico", "psicólogo", "total"]
                )

            # Encontrar linhas com "Soma" para determinar onde parar a coleta de dados
            soma_medicos_row = None
            soma_psicologos_row = None

            for row in range(3, max_row):
                if ws.cell(row=row, column=4).value == "Soma":
                    soma_medicos_row = row
                if ws.cell(row=row, column=10).value == "Soma":
                    soma_psicologos_row = row

            # Coletar dados dos médicos
            for row in range(3, max_row):
                nome_med = ws.cell(row=row, column=2).value
                # Se encontrou a linha de soma, para de coletar
                if soma_medicos_row and row >= soma_medicos_row:
                    break

                if (
                    isinstance(nome_med, str)
                    and nome_med.strip()
                    and not is_total_row(nome_med)
                ):
                    dados_medicos.append(
                        {
                            "nome": nome_med.strip(),
                            "renach": str(
                                ws.cell(row=row, column=3).value or ""
                            ).strip(),
                            "reexames": str(
                                ws.cell(row=row, column=4).value or ""
                            ).strip(),
                            "pagamento": str(
                                ws.cell(row=row, column=6).value or ""
                            ).strip(),
                        }
                    )

            # Coletar dados dos psicólogos
            for row in range(3, max_row):
                nome_psi = ws.cell(row=row, column=8).value
                # Se encontrou a linha de soma, para de coletar
                if soma_psicologos_row and row >= soma_psicologos_row:
                    break

                if (
                    isinstance(nome_psi, str)
                    and nome_psi.strip()
                    and not is_total_row(nome_psi)
                ):
                    dados_psicologos.append(
                        {
                            "nome": nome_psi.strip(),
                            "renach": str(
                                ws.cell(row=row, column=9).value or ""
                            ).strip(),
                            "reexames": str(
                                ws.cell(row=row, column=10).value or ""
                            ).strip(),
                            "pagamento": str(
                                ws.cell(row=row, column=12).value or ""
                            ).strip(),
                        }
                    )

            # Fazer uma cópia dos intervalos mesclados antes de iterar
            merged_ranges = list(ws.merged_cells.ranges)

            for merged_range in merged_ranges:
                print(f"Desmesclando: {merged_range}")
                ws.unmerge_cells(str(merged_range))  # Desfaz a mesclagem

            # Limpar planilha
            for row in range(1, max_row):
                for col in range(1, 13):
                    cell = ws.cell(row=row, column=col)
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = None
                        cell.border = borda
                        cell.font = font_regular
                        cell.alignment = alignment_center

            # Configurar cabeçalhos
            data_atual = datetime.now().strftime("%d/%m/%Y")
            usuario = (
                self.current_user
                if hasattr(self, "current_user") and self.current_user
                else "Usuário"
            )

            # Cabeçalhos principais
            ws["A1"] = f"({usuario}) Atendimento Médico {data_atual}"
            ws.merge_cells("A1:F1")
            ws["G1"] = f"({usuario}) Atendimento Psicológico {data_atual}"
            ws.merge_cells("G1:L1")

            for col in range(1, 7):
                cell = ws.cell(row=1, column=col)
                cell.font = font_bold
                cell.alignment = alignment_center

            for col in range(7, 13):
                cell = ws.cell(row=1, column=col)
                cell.font = font_bold
                cell.alignment = alignment_center

            # Cabeçalhos das colunas
            headers = ["Ordem", "Nome", "Renach", "Reexames", "Valor", "Pagamento"]
            for start_col in [1, 7]:
                for idx, header in enumerate(headers):
                    cell = ws.cell(row=2, column=start_col + idx)
                    cell.value = header
                    cell.font = font_bold
                    cell.alignment = alignment_center
                    cell.border = borda

            # Restaurar dados médicos
            for idx, dados in enumerate(dados_medicos, start=3):
                # Ordem
                ws.cell(row=idx, column=1).value = idx - 2

                # Dados do paciente
                ws.cell(row=idx, column=2).value = dados["nome"]
                ws.cell(row=idx, column=2).alignment = alignment_left

                ws.cell(row=idx, column=3).value = dados["renach"]
                ws.cell(row=idx, column=4).value = dados["reexames"]

                # Valor fixo
                valor_cell = ws.cell(row=idx, column=5)
                valor_cell.value = 148.65
                valor_cell.number_format = '"R$"#,##0.00'

                # Forma de pagamento
                ws.cell(row=idx, column=6).value = dados["pagamento"]

            # Restaurar dados psicólogos
            for idx, dados in enumerate(dados_psicologos, start=3):
                # Ordem
                ws.cell(row=idx, column=7).value = idx - 2

                # Dados do paciente
                ws.cell(row=idx, column=8).value = dados["nome"]
                ws.cell(row=idx, column=8).alignment = alignment_left

                ws.cell(row=idx, column=9).value = dados["renach"]
                ws.cell(row=idx, column=10).value = dados["reexames"]

                # Valor fixo
                valor_cell = ws.cell(row=idx, column=11)
                valor_cell.value = 192.61
                valor_cell.number_format = '"R$"#,##0.00'

                # Forma de pagamento
                ws.cell(row=idx, column=12).value = dados["pagamento"]

            # Adicionar totais médicos (uma linha abaixo do último paciente)
            if dados_medicos:
                linha_med = len(dados_medicos) + 3

                # Soma
                ws.cell(row=linha_med, column=4).value = (
                    "Soma"  # Adiciona "Soma" na coluna de reexames
                )
                ws.cell(row=linha_med, column=5).value = len(dados_medicos) * 148.65
                ws.cell(row=linha_med, column=5).number_format = '"R$"#,##0.00'

                # Médico
                ws.cell(row=linha_med + 1, column=4).value = (
                    "Médico"  # Adiciona "Médico" na coluna de reexames
                )
                ws.cell(row=linha_med + 1, column=5).value = len(dados_medicos) * 49.00
                ws.cell(row=linha_med + 1, column=5).number_format = '"R$"#,##0.00'

                # Total
                ws.cell(row=linha_med + 2, column=4).value = (
                    "Total"  # Adiciona "Total" na coluna de reexames
                )
                ws.cell(row=linha_med + 2, column=5).value = (
                    len(dados_medicos) * 148.65
                ) - (len(dados_medicos) * 49.00)
                ws.cell(row=linha_med + 2, column=5).number_format = '"R$"#,##0.00'

            # Adicionar totais psicólogos (uma linha abaixo do último paciente)
            if dados_psicologos:
                linha_psi = len(dados_psicologos) + 3

                # Soma
                ws.cell(row=linha_psi, column=10).value = (
                    "Soma"  # Adiciona "Soma" na coluna de reexames
                )
                ws.cell(row=linha_psi, column=11).value = len(dados_psicologos) * 192.61
                ws.cell(row=linha_psi, column=11).number_format = '"R$"#,##0.00'

                # Psicólogo
                ws.cell(row=linha_psi + 1, column=10).value = (
                    "Psicólogo"  # Adiciona "Psicólogo" na coluna de reexames
                )
                ws.cell(row=linha_psi + 1, column=11).value = (
                    len(dados_psicologos) * 63.50
                )
                ws.cell(row=linha_psi + 1, column=11).number_format = '"R$"#,##0.00'

                # Total
                ws.cell(row=linha_psi + 2, column=10).value = (
                    "Total"  # Adiciona "Total" na coluna de reexames
                )
                ws.cell(row=linha_psi + 2, column=11).value = (
                    len(dados_psicologos) * 192.61
                ) - (len(dados_psicologos) * 63.50)
                ws.cell(row=linha_psi + 2, column=11).number_format = '"R$"#,##0.00'

            # Ajustar largura das colunas
            larguras = {
                "A": 8,
                "B": 40,
                "C": 12,
                "D": 12,
                "E": 12,
                "F": 15,
                "G": 8,
                "H": 40,
                "I": 12,
                "J": 12,
                "K": 12,
                "L": 15,
            }
            for coluna, largura in larguras.items():
                ws.column_dimensions[coluna].width = largura

            self.planilhas.wb.save(self.file_path)
            return True

        except Exception as e:
            self.logger.error(f"Erro ao formatar planilha: {str(e)}")
            return False

    # Código de salvamento na planilha...
    def salvar_na_planilha(self, nome, renach, pagamentos, tipo_escolha):
        """
        Salva as informações do paciente na planilha Excel.
        """
        try:
            # Garantir que o workbook está carregado
            if not self.planilhas.wb:
                self.planilhas.reload_workbook()

            ws = self.planilhas.get_active_sheet()
            if not ws:
                raise Exception("Não foi possível acessar a planilha ativa")

            def encontrar_linha_insercao(coluna_inicial, coluna_soma):
                """
                Encontra a linha para inserir o novo paciente, considerando a linha de soma.
                Procura palavra 'Soma' na coluna de reexames (D para médico, J para psicólogo).
                """
                ultima_linha = 3
                linha_soma = None

                for row in range(3, ws.max_row + 1):
                    valor_nome = ws.cell(row=row, column=coluna_inicial).value
                    valor_soma = ws.cell(row=row, column=coluna_soma).value

                    if valor_soma and str(valor_soma).strip().lower() == "soma":
                        linha_soma = row
                        break
                    elif valor_nome:
                        ultima_linha = row + 1

                return linha_soma if linha_soma else ultima_linha

            alteracoes_feitas = False
            reexame_mark = "R" if self.reexame_var.get() else ""

            # Salvar na seção médica
            if tipo_escolha in ["medico", "ambos"]:
                linha_medico = encontrar_linha_insercao(2, 4)

                if ws.cell(row=linha_medico, column=4).value == "Soma":
                    ws.insert_rows(linha_medico)

                ws.cell(row=linha_medico, column=1, value=linha_medico - 2)  # Ordem
                ws.cell(row=linha_medico, column=2, value=nome)  # Nome
                ws.cell(row=linha_medico, column=3, value=renach)  # Renach
                ws.cell(row=linha_medico, column=4, value=reexame_mark)  # Reexame
                ws.cell(row=linha_medico, column=5, value=148.65)  # Valor fixo
                ws.cell(row=linha_medico, column=6, value=pagamentos)  # Pagamento

                alteracoes_feitas = True

            # Salvar na seção psicológica
            if tipo_escolha in ["psicologo", "ambos"]:
                linha_psi = encontrar_linha_insercao(8, 10)

                if ws.cell(row=linha_psi, column=10).value == "Soma":
                    ws.insert_rows(linha_psi)

                ws.cell(row=linha_psi, column=7, value=linha_psi - 2)  # Ordem
                ws.cell(row=linha_psi, column=8, value=nome)  # Nome
                ws.cell(row=linha_psi, column=9, value=renach)  # Renach
                ws.cell(row=linha_psi, column=10, value=reexame_mark)  # Reexame
                ws.cell(row=linha_psi, column=11, value=192.61)  # Valor fixo
                ws.cell(row=linha_psi, column=12, value=pagamentos)  # Pagamento

                alteracoes_feitas = True

            if alteracoes_feitas:
                self.planilhas.wb.save(self.file_path)
                return True

            return False

        except Exception as e:
            self.logger.error(f"Erro ao salvar na planilha: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao salvar na planilha: {str(e)}")
            return False

    # Código de adição de totais...
    def _adicionar_totais(
        self,
        ws,
        linha_inicio,
        col_inicio,
        valor_consulta,
        valor_profissional,
        num_pacientes,
        borda,
        font_bold,
        alignment_center,
    ):
        """
        Adiciona os totais financeiros em uma seção específica da planilha.

        Args:
            ws: Worksheet ativa
            linha_inicio (int): Linha inicial para adicionar os totais
            col_inicio (int): Coluna inicial
            valor_consulta (float): Valor da consulta
            valor_profissional (float): Valor do profissional
            num_pacientes (int): Número de pacientes
            borda: Estilo de borda a ser aplicado
            font_bold: Estilo de fonte em negrito
            alignment_center: Alinhamento centralizado
        """
        # Soma
        ws.cell(row=linha_inicio, column=col_inicio + 2, value="Soma")
        soma_cell = ws.cell(
            row=linha_inicio,
            column=col_inicio + 4,
            value=valor_consulta * num_pacientes,
        )
        soma_cell.number_format = '"R$"#,##0.00'

        # Valor profissional
        ws.cell(row=linha_inicio + 1, column=col_inicio + 2, value="Profissional")
        prof_cell = ws.cell(
            row=linha_inicio + 1,
            column=col_inicio + 4,
            value=valor_profissional * num_pacientes,
        )
        prof_cell.number_format = '"R$"#,##0.00'

        # Total
        ws.cell(row=linha_inicio + 2, column=col_inicio + 2, value="Total")
        total_cell = ws.cell(
            row=linha_inicio + 2,
            column=col_inicio + 4,
            value=(valor_consulta - valor_profissional) * num_pacientes,
        )
        total_cell.number_format = '"R$"#,##0.00'

        # Aplica formatação
        for row in range(linha_inicio, linha_inicio + 3):
            for col in range(col_inicio + 2, col_inicio + 5):
                cell = ws.cell(row=row, column=col)
                cell.border = borda
                cell.font = font_bold
                cell.alignment = alignment_center

    # Código de adição de resumo...
    def _adicionar_resumo_geral(
        self, ws, linha_inicio, num_medico, num_psi, borda, font_bold, alignment_center
    ):
        """
        Adiciona um resumo geral financeiro na planilha.

        Args:
            ws: Worksheet ativa
            linha_inicio (int): Linha inicial para adicionar o resumo
            num_medico (int): Número de atendimentos médicos
            num_psi (int): Número de atendimentos psicológicos
            borda: Estilo de borda a ser aplicado
            font_bold: Estilo de fonte em negrito
            alignment_center: Alinhamento centralizado
        """
        # Configurações
        valor_medico = 148.65
        valor_psi = 192.61
        pagamento_medico = 49.00
        pagamento_psi = 63.50

        # Cálculos
        total_medico = num_medico * valor_medico
        total_psi = num_psi * valor_psi
        total_geral = total_medico + total_psi

        pagamento_total_medico = num_medico * pagamento_medico
        pagamento_total_psi = num_psi * pagamento_psi
        total_clinica = total_geral - (pagamento_total_medico + pagamento_total_psi)

        # Lista de valores a serem adicionados
        resumo = [
            ("Atendimento Médico", total_medico),
            ("Atendimento Psicológico", total_psi),
            ("Total", total_geral),
            ("", None),  # Linha em branco
            ("Pagamento Médico", pagamento_total_medico),
            ("Pagamento Psicológico", pagamento_total_psi),
            ("Total Clínica", total_clinica),
        ]

        # Adiciona os valores
        for idx, (texto, valor) in enumerate(resumo):
            if texto:  # Pula linha em branco
                ws.cell(row=linha_inicio + idx, column=8, value=texto)
                if valor is not None:
                    valor_cell = ws.cell(row=linha_inicio + idx, column=10, value=valor)
                    valor_cell.number_format = '"R$"#,##0.00'

                # Aplica formatação
                for col in [8, 9, 10]:
                    cell = ws.cell(row=linha_inicio + idx, column=col)
                    cell.border = borda
                    cell.font = font_bold
                    cell.alignment = alignment_center

                # Mescla células para o texto
                ws.merge_cells(f"H{linha_inicio + idx}:I{linha_inicio + idx}")

    """
    SEÇÃO 6: COMUNICAÇÃO E INTEGRAÇÃO
    Esta seção contém métodos relacionados à integração com serviços externos
    """

    # Código de envio de WhatsApp...
    def enviar_whatsapp(self):
        """
        Processa o envio de mensagens via WhatsApp Web.

        Automatiza o processo de envio de mensagens com valores de atendimentos
        médicos e psicológicos para um contato ou grupo especificado.

        Utiliza Selenium WebDriver para automatizar a interação com WhatsApp Web.
        """
        janela_wpp = tk.Toplevel(self.master)
        janela_wpp.geometry("300x210")
        cor_fundo = self.master.cget("bg")
        janela_wpp.configure(bg=cor_fundo)
        self.center(janela_wpp)

        tk.Label(
            janela_wpp,
            text="Enviar para:",
            font=("Arial", 16, "bold"),
            bg=cor_fundo,
            fg="#ECF0F1",
        ).pack(anchor="center", padx=5, pady=5)

        self.wpp_entry = tk.Entry(janela_wpp)
        self.wpp_entry.pack(padx=5, pady=5)

        # Checkbutton para salvar as informações
        tk.Button(
            janela_wpp, text="Enviar", command=self.processar_envio_whatsapp
        ).pack(pady=10)

    # Código de processamento de envio...
    def processar_envio_whatsapp(self):
        """
        Processa o envio de mensagens via WhatsApp Web.

        Utiliza Selenium WebDriver para automatizar o envio de mensagens contendo
        informações sobre valores de atendimentos médicos e psicológicos.
        Realiza login automático no WhatsApp Web e envia as mensagens para o
        contato ou grupo especificado.

        Raises:
            Exception: Se houver erro na automação ou envio das mensagens
        """
        group_name = self.wpp_entry.get().strip()

        if not group_name:
            messagebox.showerror("Erro", "Insira um número, grupo ou nome")
            return

        # Preparar as informações para enviar a mensagem
        n_medico, pag_medico = self.planilhas.contar_medico()
        n_psicologo, pag_psicologo = self.planilhas.contar_psi()

        valor_medico = n_medico * 49
        valor_psicologo = n_psicologo * 63.50

        message_medico = f"Valor medico: {valor_medico}"
        message_psicologo = f"Valor psicologo: {valor_psicologo}"

        dir_path = os.getcwd()
        profile = os.path.join(dir_path, "profile", "wpp")

        # Configurar opções do Chrome
        logging.info("Configurando Chrome...")
        options = Options()
        options.add_argument(r"user-data-dir={}".format(profile))
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")

        # Inicializar o WebDriver
        logging.info("Inicializando o WebDriver...")
        service = Service(executable_path="/usr/local/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=options)

        # Acessar o WhatsApp Web
        logging.info("Abrindo WhatsApp...")
        driver.get("https://web.whatsapp.com/")

        # Aguardar até que a página esteja totalmente carregada
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//div[@role="textbox"]'))
            )
        except Exception as e:
            messagebox.showerror("Erro" f"Erro ao esnaear QR Code: {e}")
            logging.error(f"Erro ao escanear QR Code: {e}")
            driver.quit()
            return

        # Selecionar o grupo
        try:
            logging.info("Enviando mensagem...")
            barra_pesquisa = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (By.XPATH, '//*[@id="side"]/div[1]/div/div[2]/div[2]/div/div/p')
                )
            )
            barra_pesquisa.send_keys(group_name)
            time.sleep(1)
            barra_pesquisa.send_keys(Keys.ENTER)
        except Exception as e:
            messagebox.showerror(f"Erro ao selecionar grupo: {e}")
            logging.error(f"Erro ao seleconar o grupo: {e}")
            driver.quit()
            return

        # Enviar a mensagem
        try:
            logging.info("Enviando mensagem...")
            enviar_mensagem = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[1]/div/div[1]/p',
                    )
                )
            )
            enviar_mensagem.send_keys(message_medico)
            time.sleep(1)
            enviar_mensagem.send_keys(Keys.ENTER)

            enviar_mensagem = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[1]/div/div[1]/p',
                    )
                )
            )
            enviar_mensagem.send_keys(message_psicologo)
            time.sleep(1)
            enviar_mensagem.send_keys(Keys.ENTER)
            messagebox.showinfo("Mensagens enviadas")

            time.sleep(7)
        except Exception as e:
            messagebox.showerror(f"Erro ao enviar as mensagens: {e}")
        finally:
            driver.quit()

    # Código de envio de email...
    def enviar_email(self):
        """
        Configura e exibe interface para envio de email.

        Cria uma interface gráfica para coleta de informações necessárias
        para o envio de email, incluindo destinatário, assunto e opção
        para anexar arquivo XLSX.
        """
        janela_email = tk.Toplevel(self.master)
        janela_email.geometry("300x400")
        cor_fundo = self.master.cget("bg")
        janela_email.configure(bg=cor_fundo)
        self.center(janela_email)

        tk.Label(
            janela_email,
            text="Email:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_email = tk.Entry(janela_email)
        entry_email.pack(pady=5)

        tk.Label(
            janela_email,
            text="Senha:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_senha = tk.Entry(janela_email, show="*")  # Ocultar senha
        entry_senha.pack(pady=5)

        tk.Label(
            janela_email,
            text="Destinatário:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_destinatario = tk.Entry(janela_email)
        entry_destinatario.pack(pady=5)

        tk.Label(
            janela_email,
            text="Assunto:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_assunto = tk.Entry(janela_email)
        entry_assunto.pack(pady=5)

        tk.Button(
            janela_email,
            text="Selecionar XLSX",
            command=lambda: self.selecionar_xlsx(
                entry_email.get(),
                entry_senha.get(),
                entry_destinatario.get(),
                entry_assunto.get(),
            ),
        ).pack(pady=20)

    # Código de seleção de arquivo...
    def selecionar_xlsx(self, email, senha, destinatario, assunto):
        """
        Abre um diálogo para selecionar um arquivo XLSX para envio por email.

        Args:
            email (str): Email do remetente
            senha (str): Senha do email
            destinatario (str): Email do destinatário
            assunto (str): Assunto do email
        """
        if not all([email, senha, destinatario, assunto]):
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        arquivo_xlsx = filedialog.askopenfilename(
            title="Selecione o arquivo XLSX",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
        )

        if arquivo_xlsx:
            self.enviar(email, senha, destinatario, assunto, arquivo_xlsx)

    # Código de envio...
    def enviar(self, email, senha, destinatario, assunto, caminho_xlsx):
        """
        Envia um email com um arquivo XLSX anexado.

        Args:
            email (str): Email do remetente
            senha (str): Senha do email
            destinatario (str): Email do destinatário
            assunto (str): Assunto do email
            caminho_xlsx (str): Caminho do arquivo XLSX a ser anexado
        """
        smtp_server = "smtp.gmail.com"  # Para Gmail
        smtp_port = 587

        try:
            # Criando a mensagem
            msg = MIMEMultipart()
            msg["From"] = email
            msg["To"] = destinatario
            msg["Subject"] = assunto

            # Corpo do e-mail padrão
            corpo = "Segue em anexo o arquivo XLSX conforme solicitado."
            msg.attach(MIMEText(corpo, "plain"))

            # Anexar arquivo XLSX
            with open(caminho_xlsx, "rb") as arquivo:
                parte_xlsx = MIMEApplication(arquivo.read(), _subtype="xlsx")
                parte_xlsx.add_header(
                    "Content-Disposition",
                    "attachment",
                    filename=os.path.basename(caminho_xlsx),
                )
                msg.attach(parte_xlsx)

            # Contexto SSL para conexão segura
            context = ssl.create_default_context()

            # Enviando o e-mail
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls(context=context)  # Inicia a segurança TLS
                server.login(email, senha)  # Faz login no servidor
                server.send_message(msg)  # Envia a mensagem

            # Mensagem de sucesso
            messagebox.showinfo(
                "Sucesso",
                f"E-mail enviado com sucesso para {destinatario}!\nAnexo: {os.path.basename(caminho_xlsx)}",
            )

        except smtplib.SMTPAuthenticationError:
            messagebox.showerror(
                "Erro de Autenticação",
                "Verifique seu email e senha. Use uma senha de aplicativo para o Gmail.",
            )
        except Exception as e:
            messagebox.showerror("Erro ao Enviar", f"Ocorreu um erro: {str(e)}")

    """
    SEÇÃO 7: RELATÓRIOS E VISUALIZAÇÃO
    Esta seção contém métodos relacionados à geração e exibição de relatórios
    """

    # Código de exibição de totais...
    def valores_totais(self):
        """
        Exibe os valores totais e resumo financeiro.
        Cria uma janela com informações detalhadas sobre valores e quantidades.
        """
        totais = self._calcular_valores_atendimentos()
        if not totais:
            messagebox.showerror("Erro", "Não foi possível calcular os valores")
            return

        # Valores fixos
        VALORES = PaymentProcessor.SERVICE_PRICES

        # Cálculos
        dados_exibicao = {}
        for tipo in ["medico", "psicologo"]:
            n_pacientes = totais[tipo]["pacientes"]
            dados_exibicao[tipo] = {
                "pacientes": n_pacientes,
                "total": totais[tipo]["total"],
                "valor_pagar": n_pacientes * VALORES[tipo]["profissional"],
            }

        # Criação da janela
        janela_contas = tk.Toplevel(self.master)
        janela_contas.title("Contas")
        janela_contas.geometry("300x400")
        janela_contas.configure(bg="#2C3E50")

        # Função auxiliar para criar seção
        def criar_secao(titulo, dados):
            tk.Label(
                janela_contas,
                text=titulo,
                font=("Arial", 16, "bold"),
                bg="#2C3E50",
                fg="#ECF0F1",
            ).pack(pady=(15, 5))

            for label, valor in [
                ("Número de pacientes", dados["pacientes"]),
                ("Valor Total", f"R$ {dados['total']:.2f}"),
                ("Valor a Pagar", f"R$ {dados['valor_pagar']:.2f}"),
            ]:
                tk.Label(
                    janela_contas,
                    text=f"{label}: {valor}",
                    bg="#2C3E50",
                    fg="#ECF0F1",
                    font=("Arial", 12),
                ).pack(pady=2)

            tk.Label(janela_contas, text="", bg="#2C3E50").pack(pady=10)

        # Criar seções
        criar_secao("Contas Médico:", dados_exibicao["medico"])
        criar_secao("Contas Psicólogo:", dados_exibicao["psicologo"])

        # Resumo Geral
        tk.Label(
            janela_contas,
            text="Resumo Geral:",
            font=("Arial", 16, "bold"),
            bg="#2C3E50",
            fg="#ECF0F1",
        ).pack(pady=5)

        total_geral = sum(dados["total"] for dados in dados_exibicao.values())
        total_pagar = sum(dados["valor_pagar"] for dados in dados_exibicao.values())

        tk.Label(
            janela_contas,
            text=f"Total Geral: R$ {total_geral:.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(
            janela_contas,
            text=f"Total a Pagar: R$ {total_pagar:.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        self.center(janela_contas)

    def mostrar_valores_atendimentos(self):
        # Carrega os dados da planilha
        wb = self.get_active_workbook()
        ws = wb.active
        # Inicializa o total de pagamentos acumulados para cada método
        total_medico = {"Débito": 0, "Crédito": 0, "Espécie": 0, "PIX": 0}
        total_psicologo = {"Débito": 0, "Crédito": 0, "Espécie": 0, "PIX": 0}
        # Valores fixos para consulta
        VALOR_MEDICO = 148.65
        VALOR_PSICOLOGO = 192.61

        def processar_pagamento(pagamento_str, valor_padrao, totais):
            if not pagamento_str:
                return

            try:
                # Para códigos simples (D, C, E, P)
                if pagamento_str in ["D", "C", "E", "P"]:
                    metodo = self._traduzir_metodo(pagamento_str)
                    totais[metodo] += valor_padrao
                    return

                # Para formatos complexos (D:100,65|C:48,00)
                for parte in pagamento_str.split("|"):
                    if ":" not in parte:
                        continue

                    partes = parte.split(":")
                    if len(partes) != 2:
                        continue

                    metodo, valor = partes
                    metodo = self._traduzir_metodo(metodo.strip())
                    try:
                        valor = float(valor.strip().replace(",", "."))
                        totais[metodo] += valor
                    except (ValueError, KeyError):
                        continue

            except Exception as e:
                print(f"Erro ao processar pagamento '{pagamento_str}': {e}")

        # Itera sobre as linhas da planilha para calcular os valores
        for row in range(3, ws.max_row + 1):
            pagamento_medico = ws[f"F{row}"].value
            pagamento_psicologo = ws[f"L{row}"].value

            # Processa os valores de pagamento
            processar_pagamento(pagamento_medico, VALOR_MEDICO, total_medico)
            processar_pagamento(pagamento_psicologo, VALOR_PSICOLOGO, total_psicologo)

        # Criação da janela para exibir os valores
        janela_valores = tk.Toplevel(self.master)
        janela_valores.title("Valores dos Atendimentos")
        janela_valores.geometry("400x400")
        janela_valores.configure(bg="#2C3E50")

        # Exibindo valores acumulados para médico
        tk.Label(
            janela_valores,
            text="Valores - Médico:",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 20, "bold"),
        ).pack(pady=5)
        for metodo, valor in total_medico.items():
            tk.Label(
                janela_valores,
                text=f"{metodo}: R$ {valor:.2f}",
                bg="#2C3E50",
                fg="#ECF0F1",
                font=("Arial", 12, "bold"),
            ).pack()

        # Espaço entre seções
        tk.Label(janela_valores, text="", bg="#2C3E50").pack()

        # Exibindo valores acumulados para psicólogo
        tk.Label(
            janela_valores,
            text="Valores - Psicólogo:",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 20, "bold"),
        ).pack(pady=5)
        for metodo, valor in total_psicologo.items():
            tk.Label(
                janela_valores,
                text=f"{metodo}: R$ {valor:.2f}",
                bg="#2C3E50",
                fg="#ECF0F1",
                font=("Arial", 12, "bold"),
            ).pack()

        # Centraliza a janela
        self.center(janela_valores)

    def _traduzir_metodo(self, codigo):
        """Converte os códigos de pagamento em textos legíveis."""
        return {"D": "Débito", "C": "Crédito", "E": "Espécie", "P": "PIX"}.get(
            codigo, "Desconhecido"
        )

    # Código de processamento de notas...
    def processar_notas_fiscais(self):
        """Processa e emite notas fiscais com autenticação dupla."""
        driver = None
        primeira_conta = None
        segunda_conta = None

        try:
            emitir_nota = EmitirNota(self.master)

            def process_automation():
                nonlocal driver, primeira_conta, segunda_conta
                try:
                    driver = webdriver.Chrome()
                    cpfs = {"medico": [], "psicologo": [], "ambos": []}

                    # Ler o arquivo Excel
                    logging.info("Lendo o arquivo Excel")
                    df = pd.read_excel(
                        self.file_path,
                        skiprows=1,
                        sheet_name="17.10",
                        dtype={"Renach": str},
                    )

                    renach_c = df.iloc[:, 2].dropna().tolist()
                    renach_i = df.iloc[:, 8].dropna().tolist()

                    # Login no primeiro sistema (DETRAN)
                    logging.info("Acessando o site do DETRAN")
                    driver.get("https://clinicas.detran.ba.gov.br/")
                    campo_usuario = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="documento"]')
                        )
                    )

                    for numero in primeira_conta["usuario"]:
                        campo_usuario.send_keys(numero)

                    actions = ActionChains(driver)
                    actions.send_keys(Keys.TAB).perform()
                    time.sleep(1)

                    campo_senha = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="senha"]'))
                    )
                    for numero in primeira_conta["senha"]:
                        campo_senha.send_keys(numero)

                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="acessar"]'))
                    ).click()
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "/html/body/aside/section/ul/li[2]/a/span[1]")
                        )
                    ).click()
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "/html/body/aside/section/ul/li[2]/ul/li/a/span")
                        )
                    ).click()

                    # Coletar CPFs
                    barra_pesquisa = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="list_items_filter"]/label/input')
                        )
                    )

                    for dados, tipo in [(renach_c, "medico"), (renach_i, "psicologo")]:
                        for dado in dados:
                            dado = str(dado).strip()
                            barra_pesquisa.clear()
                            barra_pesquisa.send_keys(dado)
                            time.sleep(2)
                            try:
                                paciente = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located(
                                        (
                                            By.XPATH,
                                            '//*[@id="list_items"]/tbody/tr/td[3]',
                                        )
                                    )
                                )
                                cpf = paciente.text

                                if tipo == "medico" and dado in renach_i:
                                    cpfs["ambos"].append(cpf)
                                elif tipo == "medico":
                                    cpfs["medico"].append(cpf)
                                elif tipo == "psicologo" and cpf not in cpfs["ambos"]:
                                    cpfs["psicologo"].append(cpf)
                            except Exception as e:
                                logging.error(f"Erro ao coletar CPF: {e}")

                    # Filtrar CPFs duplicados
                    cpfs["medico"] = [
                        cpf for cpf in cpfs["medico"] if cpf not in cpfs["ambos"]
                    ]
                    cpfs["psicologo"] = [
                        cpf for cpf in cpfs["psicologo"] if cpf not in cpfs["ambos"]
                    ]

                    # Login no segundo sistema (NFSe)
                    logging.info("Acessando site para emissão de NTFS-e")
                    driver.get("https://nfse.salvador.ba.gov.br/")

                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="txtLogin"]')
                        )
                    ).send_keys(segunda_conta["usuario"])

                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="txtSenha"]')
                        )
                    ).send_keys(segunda_conta["senha"])

                    # Aguardar resolução do captcha
                    WebDriverWait(driver, 30).until(
                        EC.invisibility_of_element_located(
                            (By.XPATH, '//*[@id="img1"]')
                        )
                    )

                    # Emissao NFS-e
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="menu-lateral"]/li[1]/a')
                        )
                    ).click()

                    # Emitir notas para cada tipo
                    for tipo, lista_cpfs in cpfs.items():
                        for cpf in lista_cpfs:
                            try:
                                barra_pesquisa = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="tbCPFCNPJTomador"]')
                                    )
                                )
                                barra_pesquisa.clear()
                                barra_pesquisa.send_keys(cpf)
                                WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="btAvancar"]')
                                    )
                                ).click()

                                WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="ddlCNAE_chosen"]/a')
                                    )
                                ).click()

                                WebDriverWait(driver, 30).until(
                                    EC.visibility_of_element_located(
                                        (
                                            By.XPATH,
                                            '//*[@id="ddlCNAE_chosen"]/div/ul/li[2]',
                                        )
                                    )
                                ).click()

                                WebDriverWait(driver, 30).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, '//*[@id="tbAliquota"]')
                                    )
                                ).send_keys("2,5")

                                servicos = {
                                    "ambos": "Exame de sanidade física e mental",
                                    "psicologo": "Exame de sanidade mental",
                                    "medico": "Exame de sanidade física",
                                }
                                tipo_servico = servicos.get(
                                    tipo, "Exame de sanidade física"
                                )

                                WebDriverWait(driver, 30).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, '//*[@id="tbDiscriminacao"]')
                                    )
                                ).send_keys(tipo_servico)

                                valor_nota = (
                                    "148,65"
                                    if tipo == "medico"
                                    else "192,61" if tipo == "psicologo" else "341,26"
                                )
                                WebDriverWait(driver, 20).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, '//*[@id="tbValor"]')
                                    )
                                ).send_keys(valor_nota)

                                WebDriverWait(driver, 20).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="btEmitir"]')
                                    )
                                ).click()

                                # Aceitar alerta
                                WebDriverWait(driver, 20).until(EC.alert_is_present())
                                Alert(driver).accept()

                                # Voltar para emissão
                                WebDriverWait(driver, 20).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="btVoltar"]')
                                    )
                                ).click()

                                logging.info(
                                    f"Nota emitida para o CPF: {cpf}, Valor: {valor_nota}"
                                )

                            except Exception as e:
                                logging.error(f"Erro ao emitir nota: {e}")

                except Exception as e:
                    logging.error(f"Erro durante o processamento: {e}")
                    messagebox.showerror("Erro", f"Erro durante o processamento: {e}")
                finally:
                    if driver:
                        driver.quit()
                    logging.info("Processo finalizado")

            def after_first_login(usuario, senha):
                nonlocal primeira_conta
                primeira_conta = {"usuario": usuario, "senha": senha}
                self.second_window = emitir_nota.show_second_window()

                def after_second_login(result):
                    if result:
                        nonlocal segunda_conta
                        segunda_conta = {
                            "usuario": self.second_window.usuario,
                            "senha": self.second_window.senha,
                        }
                        process_automation()

                self.second_window.login_callback = after_second_login

            emitir_nota.login_callback = after_first_login
            emitir_nota.show()

        except Exception as e:
            logging.error(f"Erro ao iniciar o processo: {e}")
            messagebox.showerror("Erro", f"Erro ao iniciar o processo: {e}")
            if driver:
                driver.quit()


# Gerencia operações de banco de dados relacionadas a pacientes
class DatabaseManager:
    """Gerencia operações de banco de dados relacionadas a pacientes."""

    def __init__(self, db_path: str, logger=None):
        self.db_path = db_path
        self.logger = logger or logging.getLogger(__name__)

    def _create_history_entry(
        self, data_anterior: str, data_nova: str, status_anterior: str, status_novo: str
    ) -> dict:
        """Cria uma entrada para o histórico de alterações."""
        return {
            "data_anterior": data_anterior,
            "data_nova": data_nova,
            "status_anterior": status_anterior,
            "status_novo": status_novo,
            "atualizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def _prepare_observation(self, tipo: str, pagamentos: List[str], data: str) -> str:
        """Prepara a string de observação."""
        pagamento_info = (
            " | ".join(pagamentos) if isinstance(pagamentos, list) else str(pagamentos)
        )
        return f"Tipo: {tipo}\nPagamento: {pagamento_info}\nRegistrado em: {data}"

    def _update_patient(
        self,
        cursor: sqlite3.Cursor,
        nome: str,
        renach: str,
        data_atual: str,
        status: str,
        observacao: str,
    ) -> None:
        """Atualiza os dados de um paciente existente."""
        cursor.execute(
            """
            UPDATE marcacoes 
            SET nome = ?, data_agendamento = ?, status_comparecimento = ?, observacao = ?
            WHERE renach = ?
        """,
            (nome, data_atual, status, observacao, renach),
        )

    def _insert_new_patient(
        self,
        cursor: sqlite3.Cursor,
        nome: str,
        renach: str,
        data_atual: str,
        status: str,
        observacao: str,
    ) -> None:
        """Insere um novo paciente no banco de dados."""
        cursor.execute(
            """
            INSERT INTO marcacoes (
                nome, renach, telefone, data_agendamento, 
                status_comparecimento, observacao, historico_comparecimento
            ) VALUES (?, ?, ?, ?, ?, ?, '[]')
        """,
            (nome, renach, "", data_atual, status, observacao),
        )

    def adicionar_paciente(
        self, nome: str, renach: str, pagamentos: List[str], tipo: str
    ) -> bool:
        """
        Adiciona ou atualiza paciente no banco de dados.

        Args:
            nome: Nome do paciente
            renach: Número do RENACH
            pagamentos: Lista de strings com informações de pagamento
            tipo: Tipo de atendimento

        Returns:
            bool: True se operação foi bem sucedida, False caso contrário
        """
        try:
            with DatabaseConnection(self.db_path) as conn:
                cursor = conn.cursor()
                data_atual = datetime.now().strftime("%Y-%m-%d")
                STATUS_COMPARECEU = "attended"
                observacao = self._prepare_observation(tipo, pagamentos, data_atual)

                # Verifica se paciente existe
                cursor.execute(
                    """
                    SELECT data_agendamento, status_comparecimento, historico_comparecimento
                    FROM marcacoes 
                    WHERE renach = ?
                    ORDER BY data_agendamento DESC 
                    LIMIT 1
                """,
                    (renach,),
                )

                resultado = cursor.fetchone()

                if resultado:
                    data_anterior, status_anterior, historico_atual = resultado

                    # Atualiza dados do paciente
                    self._update_patient(
                        cursor, nome, renach, data_atual, STATUS_COMPARECEU, observacao
                    )

                    # Atualiza histórico
                    try:
                        historico = (
                            json.loads(historico_atual) if historico_atual else []
                        )
                    except json.JSONDecodeError:
                        historico = []

                    historico.append(
                        self._create_history_entry(
                            data_anterior,
                            data_atual,
                            status_anterior,
                            STATUS_COMPARECEU,
                        )
                    )

                    cursor.execute(
                        """
                        UPDATE marcacoes 
                        SET historico_comparecimento = ?
                        WHERE renach = ?
                    """,
                        (json.dumps(historico), renach),
                    )

                    self.logger.info(f"Atualizado status do paciente {renach}")
                    mensagem = "Paciente atualizado com status Compareceu!"

                else:
                    # Insere novo paciente
                    self._insert_new_patient(
                        cursor, nome, renach, data_atual, STATUS_COMPARECEU, observacao
                    )

                    self.logger.info(f"Novo paciente {renach} adicionado")
                    mensagem = "Novo paciente adicionado com status Compareceu!"

                conn.commit()
                messagebox.showinfo("Sucesso", mensagem)
                return True

        except sqlite3.Error as e:
            self.logger.error(f"Erro no banco de dados: {e}")
            messagebox.showerror(
                "Erro", f"Erro ao processar operação no banco de dados: {e}"
            )
            return False
        except Exception as e:
            self.logger.error(f"Erro inesperado: {e}")
            messagebox.showerror("Erro", f"Erro inesperado: {e}")
            return False


# Classe para processamento centralizado de pagamentos
class PaymentProcessor:
    """Classe para processamento centralizado de pagamentos."""

    PAYMENT_TYPES = {"D": "Débito", "C": "Crédito", "E": "Espécie", "P": "PIX"}

    SERVICE_PRICES = {
        "medico": {"consulta": 148.65, "profissional": 49.00},
        "psicologo": {"consulta": 192.61, "profissional": 63.50},
        "ambos": {"consulta": 341.26, "profissional": 112.50},
    }

    @staticmethod
    def convert_currency_value(value_str: str) -> float:
        """
        Converte string de valor monetário para float.

        Args:
            value_str: String contendo valor monetário (ex: "R$ 148,65" ou "148.65")

        Returns:
            float: Valor convertido ou None se inválido

        Raises:
            ValueError: Se o valor não puder ser convertido
        """
        if not value_str:
            return 0.0

        try:
            # Remove R$, espaços e substitui vírgula por ponto
            clean_value = value_str.replace("R$", "").replace(" ", "").replace(",", ".")
            return float(clean_value)
        except ValueError:
            raise ValueError(f"Valor inválido para conversão: {value_str}")

    @classmethod
    def format_currency(cls, value: float) -> str:
        """
        Formata valor float para string monetária.

        Args:
            value: Valor numérico a ser formatado

        Returns:
            str: Valor formatado (ex: "148,65")
        """
        return f"{value:.2f}".replace(".", ",")

    @classmethod
    def calculate_service_value(cls, service_type: str) -> float:
        """
        Calcula valor do serviço baseado no tipo.

        Args:
            service_type: Tipo de serviço ('medico', 'psicologo' ou 'ambos')

        Returns:
            float: Valor total do serviço

        Raises:
            ValueError: Se o tipo de serviço for inválido
        """
        if service_type not in cls.SERVICE_PRICES:
            raise ValueError(f"Tipo de serviço inválido: {service_type}")
        return cls.SERVICE_PRICES[service_type]["consulta"]

    @classmethod
    def validate_payment_total(cls, payments: dict, expected_total: float) -> bool:
        """
        Valida se o total dos pagamentos corresponde ao valor esperado.

        Args:
            payments: Dicionário com valores por forma de pagamento
            expected_total: Valor total esperado

        Returns:
            bool: True se válido, False caso contrário

        Raises:
            ValueError: Se houver erro na validação
        """
        try:
            total = sum(
                cls.convert_currency_value(value)
                for value in payments.values()
                if value
            )
            return (
                abs(total - expected_total) < 0.01
            )  # Permite pequena diferença por arredondamento
        except ValueError as e:
            raise ValueError(f"Erro ao validar pagamentos: {str(e)}")

    @classmethod
    def process_payment_methods(cls, payment_data: dict) -> str:
        """
        Processa e formata métodos de pagamento.

        Se houver apenas um método, retorna apenas o código (D, C, E ou P).
        Se houver múltiplos métodos, retorna no formato "D:3000|E:4127".

        Args:
            payment_data: Dicionário com códigos de pagamento como chaves
                        e valores como strings (ex: {"D": "148,65", "C": ""})

        Returns:
            str: String formatada com os métodos de pagamento
                - Um método: apenas o código (ex: "D")
                - Múltiplos métodos: códigos e valores sem formatação (ex: "D:3000|E:4127")

        Examples:
            >>> PaymentProcessor.process_payment_methods({"D": "148,65", "C": "", "E": "", "P": ""})
            "D"
            >>> PaymentProcessor.process_payment_methods({"D": "30,00", "E": "41,27"})
            "D:3000|E:4127"
        """
        try:
            # Filtra apenas os métodos selecionados (com valores não vazios)
            selected_payments = {
                code: value
                for code, value in payment_data.items()
                if value and str(value).strip()
            }

            # Se houver apenas um método de pagamento, retorna só o código
            if len(selected_payments) == 1:
                return list(selected_payments.keys())[0]

            # Se houver múltiplos métodos, retorna no formato "D:3000|E:4127"
            formatted_parts = []
            for code, value in selected_payments.items():
                # Converte o valor para float e formata com 2 casas decimais
                value_float = cls.convert_currency_value(value)
                formatted_value = f"{value_float:.2f}".replace(".", ",")
                formatted_parts.append(f"{code}:{formatted_value}")

            return "|".join(formatted_parts)

        except Exception as e:
            raise ValueError(f"Erro ao processar métodos de pagamento: {str(e)}")

    @classmethod
    def calculate_professional_payment(
        cls, service_type: str, num_patients: int
    ) -> float:
        """
        Calcula pagamento do profissional baseado no tipo de serviço e número de pacientes.

        Args:
            service_type: Tipo de serviço ('medico', 'psicologo' ou 'ambos')
            num_patients: Número de pacientes atendidos

        Returns:
            float: Valor total a ser pago ao profissional

        Raises:
            ValueError: Se o tipo de serviço for inválido
        """
        if service_type not in cls.SERVICE_PRICES:
            raise ValueError(f"Tipo de serviço inválido: {service_type}")

        return cls.SERVICE_PRICES[service_type]["profissional"] * num_patients

    @classmethod
    def parse_payment_string(cls, payment_str: str) -> dict:
        """
        Converte uma string de pagamento no formato 'E:300,00|P:41,27' para dicionário.

        Args:
            payment_str: String com informações de pagamento

        Returns:
            dict: Dicionário com valores de pagamento por método

        Raises:
            ValueError: Se a string de pagamento estiver em formato inválido
        """
        try:
            result = {}
            # Se for apenas um código
            if len(payment_str) == 1 and payment_str in cls.PAYMENT_TYPES:
                return {payment_str: cls.PAYMENT_TYPES[payment_str]}

            # Se for múltiplos pagamentos
            for part in payment_str.split("|"):
                if ":" in part:
                    code, value = part.split(":")
                    if code in cls.PAYMENT_TYPES:
                        result[code] = value
            return result
        except Exception as e:
            raise ValueError(f"Erro ao processar string de pagamento: {str(e)}")

    @classmethod
    def get_payment_description(cls, payment_str: str) -> str:
        """
        Gera descrição legível dos pagamentos.

        Args:
            payment_str: String com informações de pagamento

        Returns:
            str: Descrição formatada dos pagamentos

        Example:
            "E:300,00|P:41,27" -> "Espécie: R$ 300,00, PIX: R$ 41,27"
            "D" -> "Débito"
        """
        try:
            # Se for apenas um código
            if len(payment_str) == 1 and payment_str in cls.PAYMENT_TYPES:
                return cls.PAYMENT_TYPES[payment_str]

            # Se for múltiplos pagamentos
            parts = []
            for part in payment_str.split("|"):
                if ":" in part:
                    code, value = part.split(":")
                    if code in cls.PAYMENT_TYPES:
                        parts.append(f"{cls.PAYMENT_TYPES[code]}: R$ {value}")
            return ", ".join(parts)
        except Exception as e:
            return f"Erro ao processar pagamento: {str(e)}"


class GerenciadorPlanilhas:
    """
    SEÇÃO 1: INICIALIZAÇÃO E CONFIGURAÇÃO
    """

    # Inicializa o gerenciador com janela principal e sistema de contas
    def __init__(self, master, sistema_contas):
        self.master = master
        self.sistema_contas = sistema_contas
        self.file_path = None
        self.sheet_name = None
        self.active_window = None

        # Create instance of ConfigManager
        self.config_manager = ConfigManager()

        # Get configs using instance method
        self.config = self.config_manager.get_config("UI_CONFIG")
        self.app_config = self.config_manager.get_config("APP_CONFIG")

    """
    SEÇÃO 2: INTERFACE GRÁFICA
    """

    # Abre janela principal do gerenciador
    def abrir_gerenciador(self):
        if self.active_window:
            self.active_window.lift()
            return

        self.active_window = Toplevel(self.master)
        self.active_window.title("Gerenciador de Planilhas")

        # Use window dimensions from config
        window_width = self.app_config["window"]["min_width"]
        window_height = self.app_config["window"]["min_height"]
        screen_width = self.active_window.winfo_screenwidth()
        screen_height = self.active_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        self.active_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.active_window.minsize(
            self.app_config["window"]["min_width"],
            self.app_config["window"]["min_height"],
        )
        self.active_window.maxsize(
            self.app_config["window"]["max_width"],
            self.app_config["window"]["max_height"],
        )

        # Configure window background
        self.active_window.configure(bg=self.config["colors"]["background"])

        self.active_window.grid_columnconfigure(0, weight=1)
        self.active_window.grid_rowconfigure(0, weight=1)

        self._setup_interface()
        self.active_window.protocol("WM_DELETE_WINDOW", self._on_closing)
        self.active_window.transient(self.master)
        self.active_window.grab_set()

    # Configura os elementos da interface
    def _setup_interface(self):
        style = ttk.Style()
        style.configure("Custom.TFrame", background=self.config["colors"]["background"])
        style.configure("Custom.TLabelframe", background=self.config["colors"]["frame"])
        style.configure(
            "Custom.TLabelframe.Label",
            foreground=self.config["colors"]["text"],
            background=self.config["colors"]["frame"],
        )
        style.configure(
            "Custom.TButton",
            background=self.config["colors"]["button"],
            foreground=self.config["colors"]["text"],
        )
        style.configure(
            "Custom.TEntry",
            fieldbackground=self.config["colors"]["frame"],
            foreground=self.config["colors"]["text"],
        )

        main_frame = ttk.Frame(self.active_window, padding=20, style="Custom.TFrame")
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_columnconfigure(0, weight=1)

        self.active_window.configure(bg=self.config["colors"]["background"])

        title_label = ttk.Label(
            main_frame,
            text="Gerenciador de Planilhas Excel",
            font=("Segoe UI", 18, "bold"),
            foreground=self.config["colors"]["title"],
            background=self.config["colors"]["background"],
        )
        title_label.grid(row=0, column=0, pady=(0, 20))

        file_frame = ttk.LabelFrame(
            main_frame, text="Arquivo Atual", padding=10, style="Custom.TLabelframe"
        )
        file_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        file_frame.grid_columnconfigure(0, weight=1)

        self.lbl_arquivo = ttk.Label(
            file_frame,
            text=(
                self.sistema_contas.file_path
                if hasattr(self.sistema_contas, "file_path")
                else "Nenhum arquivo selecionado"
            ),
            wraplength=500,
            foreground=self.config["colors"]["text"],
            background=self.config["colors"]["frame"],
        )
        self.lbl_arquivo.grid(row=0, column=0, sticky="ew", padx=5)

        list_frame = ttk.LabelFrame(
            main_frame,
            text="Planilhas Disponíveis",
            padding=10,
            style="Custom.TLabelframe",
        )
        list_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 20))
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(0, weight=1)

        self.listbox = Listbox(
            list_frame,
            font=("Segoe UI", 10),
            selectmode=SINGLE,
            height=10,
            bg=self.config["colors"]["frame"],
            fg=self.config["colors"]["text"],
            selectbackground=self.config["colors"]["button"],
            selectforeground=self.config["colors"]["text"],
            borderwidth=1,
            relief="solid",
        )
        self.listbox.grid(row=0, column=0, sticky="nsew")

        create_frame = ttk.LabelFrame(
            main_frame, text="Criar Nova Sheet", padding=10, style="Custom.TLabelframe"
        )
        create_frame.grid(row=3, column=0, sticky="ew", pady=(0, 20))
        create_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(
            create_frame,
            text="Nome:",
            font=("Segoe UI", 10),
            foreground=self.config["colors"]["text"],
            background=self.config["colors"]["frame"],
        ).grid(row=0, column=0, padx=(0, 10), sticky="w")

        self.nova_sheet_entry = ttk.Entry(create_frame, style="Custom.TEntry")
        self.nova_sheet_entry.grid(row=0, column=1, sticky="ew")

        button_frame = ttk.Frame(main_frame, style="Custom.TFrame")
        button_frame.grid(row=4, column=0, sticky="ew")
        for i in range(2):
            button_frame.grid_columnconfigure(i, weight=1)

        buttons = [
            ("Nova Planilha Excel", self.criar_nova_planilha),
            ("Abrir Planilha Existente", self.abrir_planilha),
            ("Selecionar Sheet", self.selecionar_sheet),
            ("Criar Nova Sheet", self.criar_nova_sheet),
        ]

        for idx, (text, command) in enumerate(buttons):
            ttk.Button(
                button_frame, text=text, command=command, style="Custom.TButton"
            ).grid(row=idx // 2, column=idx % 2, padx=5, pady=5, sticky="ew")

        self.atualizar_lista_sheets()

    # Fecha a janela do gerenciador
    def _on_closing(self):
        """Handler para quando a janela for fechada"""
        self.active_window.destroy()
        self.active_window = None

    """
    SEÇÃO 3: MANIPULAÇÃO DE ARQUIVOS
    """

    # Cria nova planilha Excel
    def criar_nova_planilha(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=self.app_config["file_types"]
        )

    # Abre planilha Excel existente
    def abrir_planilha(self):
        """Abre uma planilha Excel existente"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if file_path:
            try:
                wb = load_workbook(file_path)
                self.sistema_contas.file_path = file_path

                # Pega a sheet ativa atual
                active_sheet = wb.active
                self.sistema_contas.sheet_name = active_sheet.title

                wb.close()
                self.lbl_arquivo.config(text=file_path)
                self.atualizar_lista_sheets()
                messagebox.showinfo(
                    "Sucesso",
                    f"Planilha aberta com sucesso! Sheet ativa: {active_sheet.title}",
                )
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir planilha: {str(e)}")

    """
    SEÇÃO 4: GERENCIAMENTO DE SHEETS
    """

    # Atualiza lista de sheets disponíveis
    def atualizar_lista_sheets(self):
        """Atualiza a lista de sheets disponíveis"""
        self.listbox.delete(0, END)
        if (
            hasattr(self.sistema_contas, "file_path")
            and self.sistema_contas.file_path
            and os.path.exists(self.sistema_contas.file_path)
        ):
            try:
                wb = load_workbook(self.sistema_contas.file_path)
                for sheet in wb.sheetnames:
                    self.listbox.insert(END, sheet)
                wb.close()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao listar planilhas: {str(e)}")

    # Seleciona e ativa uma sheet
    def selecionar_sheet(self):
        """Seleciona uma sheet existente e a torna ativa"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showerror("Erro", "Selecione uma planilha!")
            return

        nome_sheet = self.listbox.get(selection[0])
        try:
            wb = load_workbook(self.sistema_contas.file_path)
            if nome_sheet in wb.sheetnames:
                # Define a sheet selecionada como ativa
                wb.active = wb[nome_sheet]
                wb.save(self.sistema_contas.file_path)

                # Atualiza o nome da sheet no sistema_contas
                self.sistema_contas.sheet_name = nome_sheet

                wb.close()
                messagebox.showinfo(
                    "Sucesso", f"Planilha '{nome_sheet}' selecionada e ativada!"
                )
                self.active_window.destroy()
                self.active_window = None
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao selecionar planilha: {str(e)}")

    # Cria nova sheet e a torna ativa
    def criar_nova_sheet(self):
        """Cria uma nova sheet e a torna ativa"""
        nome_sheet = self.nova_sheet_entry.get().strip()
        if not nome_sheet:
            messagebox.showerror("Erro", "Digite um nome para a nova planilha!")
            return

        if (
            not hasattr(self.sistema_contas, "file_path")
            or not self.sistema_contas.file_path
        ):
            messagebox.showerror("Erro", "Primeiro abra ou crie uma planilha Excel!")
            return

        try:
            wb = load_workbook(self.sistema_contas.file_path)
            if nome_sheet in wb.sheetnames:
                messagebox.showerror("Erro", "Já existe uma planilha com este nome!")
                wb.close()
                return

            # Cria nova sheet e a torna ativa
            new_sheet = wb.create_sheet(title=nome_sheet)
            wb.active = new_sheet
            wb.save(self.sistema_contas.file_path)
            wb.close()

            # Atualiza o nome da sheet no sistema_contas
            self.sistema_contas.sheet_name = nome_sheet

            self.atualizar_lista_sheets()
            messagebox.showinfo(
                "Sucesso", f"Planilha '{nome_sheet}' criada e ativada com sucesso!"
            )
            self.active_window.destroy()
            self.active_window = None
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar planilha: {str(e)}")


@dataclass
class PatientData:
    """Estrutura de dados imutável para informações do paciente."""

    nome: str
    renach: str
    pagamento: str
    tipo: str
    search_text: str


# Classe otimizada para exibição de informações de pacientes
class PatientInfoDisplay:
    """Classe otimizada para exibição de informações de pacientes."""

    def __init__(self, master: tk.Tk, planilhas, logger=None):
        self.master = master
        self.planilhas = planilhas
        self.logger = logger or logging.getLogger(__name__)
        self.config_manager = ConfigManager()

        # Sistema de cache melhorado com TTL
        self.data_cache = {
            "medico": [],
            "psi": [],
            "last_update": None,
            "last_filters": {},
            "timer": None,
            "ttl": 300,  # 5 minutos de TTL para o cache
        }

        # Obter cores do ConfigManager
        ui_config = self.config_manager.get_config("UI_CONFIG")
        self.theme = {
            "background": ui_config["colors"]["background"],
            "secondary_bg": ui_config["colors"]["frame"],
            "text": ui_config["colors"]["text"],
            "accent": ui_config["colors"]["button"],
            "header": ui_config["colors"]["frame"],
            "highlight": ui_config["colors"]["border"],
            "separator": ui_config["colors"]["border"],
            "hover": ui_config["colors"]["button_hover"],
            "error": "#e74c3c",  # Mantido para consistência
        }

        # Configurações de UI responsiva
        self.ui_config = {
            "min_width": 800,
            "min_height": 600,
            "padding": ui_config["padding"]["default"],
            "animation_duration": 200,
        }

        # Referências de UI com estado
        self.ui_refs = {}

        # Estado de ordenação
        self.sort_state = {"column": None, "reverse": False}

    @lru_cache(maxsize=1000)
    def _process_payment(self, value: str) -> str:
        """Processa e formata valores de pagamento com cache."""
        if not value:
            return ""
        try:
            if isinstance(value, (int, float)):
                return f"R$ {float(value):.2f}"
            value_str = str(value).replace("R$", "").replace(",", ".").strip()
            return f"R$ {float(value_str):.2f}"
        except:
            return str(value).strip()

    def _load_data(self) -> bool:
        """Carrega e processa os dados da planilha com cache inteligente."""
        try:
            current_time = time.time()

            # Verifica se o cache ainda é válido
            if (
                self.data_cache["last_update"]
                and current_time - self.data_cache["last_update"]
                < self.data_cache["ttl"]
            ):
                return True

            self.planilhas.reload_workbook()
            wb = self.planilhas.wb

            if not wb:
                return False

            ws = wb.active
            if not ws:
                messagebox.showerror("Erro", "Planilha inválida")
                return False

            # Processamento em lote otimizado
            med_data = []
            psi_data = []

            # Pré-aloca as listas para melhor performance
            max_rows = ws.max_row
            med_data = []
            psi_data = []

            # Processa em chunks para melhor performance
            chunk_size = 100
            for start_row in range(3, max_rows + 1, chunk_size):
                end_row = min(start_row + chunk_size, max_rows + 1)

                rows = list(ws.iter_rows(min_row=start_row, max_row=end_row))

                for row in rows:
                    # Processamento de médicos
                    if row[1].value:
                        nome = str(row[1].value).strip().upper()
                        if not any(
                            x in nome.lower()
                            for x in ["soma", "médico", "psicólogo", "total"]
                        ):
                            med_data.append(
                                PatientData(
                                    nome=nome,
                                    renach=str(row[2].value or "").strip(),
                                    pagamento=self._process_payment(row[5].value),
                                    tipo="Médico",
                                    search_text=f"{nome.lower()} {str(row[2].value or '').lower()}",
                                )
                            )

                    # Processamento de psicólogos
                    if len(row) > 7 and row[7].value:
                        nome = str(row[7].value).strip().upper()
                        if not any(
                            x in nome.lower()
                            for x in ["soma", "médico", "psicólogo", "total"]
                        ):
                            psi_data.append(
                                PatientData(
                                    nome=nome,
                                    renach=str(row[8].value or "").strip(),
                                    pagamento=self._process_payment(row[11].value),
                                    tipo="Psicólogo",
                                    search_text=f"{nome.lower()} {str(row[8].value or '').lower()}",
                                )
                            )

            # Atualiza o cache
            self.data_cache.update(
                {"medico": med_data, "psi": psi_data, "last_update": current_time}
            )

            return bool(med_data or psi_data)

        except Exception as e:
            self.logger.error(f"Erro ao carregar dados: {e}")
            return False

    def _create_ui(self) -> Tuple[tk.Toplevel, Dict]:
        """Cria e retorna a interface do usuário."""
        window = tk.Toplevel(self.master)
        window.title("Informações dos Pacientes")
        window.geometry("1200x800")
        window.configure(bg=self.theme["background"])

        # Frames principais
        frames = self._create_frames(window)

        # Controles de filtro
        filters = self._create_filters(frames["control"])

        # Tabela
        table = self._create_table(frames["table"])

        # Barra de status
        stats_label = tk.Label(
            frames["stats"],
            bg=self.theme["background"],
            fg=self.theme["text"],
            font=("Arial", 10, "bold"),
        )
        stats_label.pack(pady=5)

        self.ui_refs = {
            "window": window,
            "frames": frames,
            "filters": filters,
            "table": table,
            "stats": stats_label,
        }

        return window, self.ui_refs

    def _create_frames(self, window: tk.Toplevel) -> Dict[str, tk.Frame]:
        """Cria e retorna os frames principais."""
        frames = {
            "main": tk.Frame(window, bg=self.theme["background"]),
            "control": tk.Frame(window, bg=self.theme["background"]),
            "table": tk.Frame(window, bg=self.theme["background"]),
            "stats": tk.Frame(window, bg=self.theme["background"]),
        }

        frames["main"].pack(fill="both", expand=True, padx=20, pady=10)
        frames["control"].pack(in_=frames["main"], fill="x", pady=(0, 10))
        frames["table"].pack(in_=frames["main"], fill="both", expand=True)
        frames["stats"].pack(in_=frames["main"], fill="x", pady=10)

        return frames

    def _create_filters(self, parent: tk.Frame) -> Dict[str, tk.Variable]:
        """Cria e retorna os controles de filtro."""
        filters = {
            "type": tk.StringVar(value="todos"),
            "search": tk.StringVar(),
            "payment": tk.StringVar(),
        }

        filter_frame = tk.Frame(parent, bg=self.theme["background"])
        filter_frame.pack(fill="x", padx=5)

        # Tipo de atendimento
        type_frame = self._create_filter_section(filter_frame, "Tipo de Atendimento")
        options = [("todos", "Todos"), ("medico", "Médico"), ("psi", "Psicólogo")]
        for value, text in options:
            tk.Radiobutton(
                type_frame,
                text=text,
                variable=filters["type"],
                value=value,
                bg=self.theme["background"],
                fg=self.theme["text"],
                selectcolor=self.theme["header"],
                command=lambda: self._delayed_filter(),
            ).pack(side="left", padx=5)

        # Busca
        search_frame = self._create_filter_section(
            filter_frame, "Buscar por Nome/RENACH"
        )
        tk.Entry(search_frame, textvariable=filters["search"], width=30).pack(
            padx=5, pady=2
        )

        # Pagamento
        payment_frame = self._create_filter_section(
            filter_frame, "Filtrar por Forma de Pagamento"
        )
        tk.Entry(payment_frame, textvariable=filters["payment"], width=20).pack(
            padx=5, pady=2
        )

        for var in filters.values():
            var.trace("w", lambda *args: self._delayed_filter())

        return filters

    def _create_filter_section(self, parent: tk.Frame, title: str) -> tk.LabelFrame:
        """Cria uma seção de filtro com título."""
        frame = tk.LabelFrame(
            parent,
            text=title,
            bg=self.theme["background"],
            fg=self.theme["text"],
            font=self.config_manager.get_config("UI_CONFIG")["fonts"]["normal"],
        )
        frame.pack(side="left", padx=5, pady=5)
        return frame

    def _create_table(self, parent: tk.Frame) -> Dict:
        """Cria uma tabela moderna e responsiva."""
        # Frame principal com bordas arredondadas
        table_container = tk.Frame(
            parent,
            bg=self.theme["background"],
            highlightbackground=self.theme["accent"],
            highlightthickness=1,
        )
        table_container.pack(fill="both", expand=True, padx=5, pady=5)

        # Canvas com scrollbar suave
        canvas = tk.Canvas(
            table_container, bg=self.theme["background"], highlightthickness=0
        )

        scrollbar = ttk.Scrollbar(
            table_container, orient="vertical", command=canvas.yview
        )

        table_frame = tk.Frame(canvas, bg=self.theme["background"])

        # Configuração do scroll
        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _bound_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbound_to_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        def _on_mousewheel(event):
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                if event.delta > 0:
                    canvas.yview_scroll(-1, "units")
                else:
                    canvas.yview_scroll(1, "units")

        table_frame.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Enter>", _bound_to_mousewheel)
        canvas.bind("<Leave>", _unbound_to_mousewheel)

        # Cabeçalhos clicáveis para ordenação
        headers = [
            ("Nº", 5),
            ("Nome", 30),
            ("RENACH", 10),
            ("Forma de Pagamento", 20),
            ("Tipo", 10),
        ]

        fonts = self.config_manager.get_config("UI_CONFIG")["fonts"]
        for col, (header, width) in enumerate(headers):
            header_frame = tk.Frame(table_frame, bg=self.theme["header"])
            header_frame.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

            label = tk.Label(
                header_frame,
                text=header,
                bg=self.theme["header"],
                fg=self.theme["text"],
                font=fonts["header"],
                padx=10,
                pady=8,
            )
            label.pack(fill="both", expand=True)
            label.bind("<Button-1>", lambda e, col=col: self._sort_table(col))
            label.bind(
                "<Enter>", lambda e, widget=label: self._on_header_hover(widget, True)
            )
            label.bind(
                "<Leave>", lambda e, widget=label: self._on_header_hover(widget, False)
            )

        canvas.create_window((0, 0), window=table_frame, anchor="nw", tags=("table",))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig("table", width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for i in range(5):
            table_frame.grid_columnconfigure(i, weight=1)

        return {
            "container": table_container,
            "frame": table_frame,
            "canvas": canvas,
            "scrollbar": scrollbar,
        }

    def _create_table(self, parent: tk.Frame) -> Dict:
        """Cria uma tabela moderna e responsiva."""
        # Frame principal com bordas arredondadas
        table_container = tk.Frame(
            parent,
            bg=self.theme["background"],
            highlightbackground=self.theme["accent"],
            highlightthickness=1,
        )
        table_container.pack(fill="both", expand=True, padx=5, pady=5)

        # Canvas com scrollbar suave
        canvas = tk.Canvas(
            table_container, bg=self.theme["background"], highlightthickness=0
        )

        scrollbar = ttk.Scrollbar(
            table_container, orient="vertical", command=canvas.yview
        )

        table_frame = tk.Frame(canvas, bg=self.theme["background"])

        # Configuração do scroll
        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _bound_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel)
            canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbound_to_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        def _on_mousewheel(event):
            # Windows
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
            else:
                # Tratamento do evento para Windows
                if event.delta > 0:
                    canvas.yview_scroll(-1, "units")
                else:
                    canvas.yview_scroll(1, "units")

        table_frame.bind("<Configure>", _on_frame_configure)

        # Vincula os eventos de scroll
        canvas.bind("<Enter>", _bound_to_mousewheel)
        canvas.bind("<Leave>", _unbound_to_mousewheel)

        # Cabeçalhos clicáveis para ordenação
        headers = [
            ("Nº", 5),
            ("Nome", 30),
            ("RENACH", 10),
            ("Forma de Pagamento", 20),
            ("Tipo", 10),
        ]

        for col, (header, width) in enumerate(headers):
            header_frame = tk.Frame(table_frame, bg=self.theme["header"])
            header_frame.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

            label = tk.Label(
                header_frame,
                text=header,
                bg=self.theme["header"],
                fg=self.theme["text"],
                font=("Arial", 11, "bold"),
                padx=10,
                pady=8,
            )
            label.pack(fill="both", expand=True)

            # Adiciona funcionalidade de ordenação
            label.bind("<Button-1>", lambda e, col=col: self._sort_table(col))
            label.bind(
                "<Enter>", lambda e, widget=label: self._on_header_hover(widget, True)
            )
            label.bind(
                "<Leave>", lambda e, widget=label: self._on_header_hover(widget, False)
            )

        # Configuração do canvas e scrollbar
        canvas.create_window((0, 0), window=table_frame, anchor="nw", tags=("table",))

        def _on_canvas_configure(event):
            canvas.itemconfig("table", width=event.width)

        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Layout responsivo - alterada a ordem de empacotamento
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Configurações de grid
        for i in range(5):
            table_frame.grid_columnconfigure(i, weight=1)

        return {
            "container": table_container,
            "frame": table_frame,
            "canvas": canvas,
            "scrollbar": scrollbar,
        }

    def _smooth_scroll(self, scrollbar):
        """Implementa scrolling suave."""

        def scroll(*args):
            scrollbar.set(*args)
            self.ui_refs["table"]["canvas"].yview_moveto(args[0])

        return scroll

    def _sort_table(self, column: int):
        """Ordena a tabela pela coluna clicada."""
        if self.sort_state["column"] == column:
            self.sort_state["reverse"] = not self.sort_state["reverse"]
        else:
            self.sort_state["column"] = column
            self.sort_state["reverse"] = False

        self._apply_filters()

    def _on_header_hover(self, widget: tk.Label, entering: bool):
        """Efeito hover nos cabeçalhos."""
        widget.configure(bg=self.theme["hover"] if entering else self.theme["header"])

    def _update_table(self, data: List[PatientData]) -> None:
        """Atualiza a tabela com os dados filtrados."""
        table = self.ui_refs["table"]

        # Limpa tabela preservando cabeçalho
        for widget in table["frame"].winfo_children():
            if int(widget.grid_info()["row"]) > 0:
                widget.destroy()

        fonts = self.config_manager.get_config("UI_CONFIG")["fonts"]
        medicos = [p for p in data if p.tipo == "Médico"]
        psicologos = [p for p in data if p.tipo == "Psicólogo"]

        row = 2
        for lista, tipo in [(medicos, "Médico"), (psicologos, "Psicólogo")]:
            for idx, patient in enumerate(lista, 1):
                bg_color = (
                    self.theme["highlight"]
                    if idx % 2 == 0
                    else self.theme["background"]
                )

                cells = [
                    (str(idx), "center", 5),
                    (patient.nome, "w", 30),
                    (patient.renach, "center", 10),
                    (patient.pagamento, "w", 20),
                    (patient.tipo, "center", 10),
                ]

                for col, (text, anchor, width) in enumerate(cells):
                    tk.Label(
                        table["frame"],
                        text=text,
                        bg=bg_color,
                        fg=self.theme["text"],
                        font=fonts["normal"],
                        anchor=anchor,
                        width=width,
                        padx=10,
                        pady=5,
                    ).grid(row=row, column=col, sticky="nsew", padx=1, pady=1)

                row += 1

            # Adiciona separador entre médicos e psicólogos
            if tipo == "Médico" and psicologos:
                separator = tk.Frame(
                    table["frame"], height=2, bg=self.theme["separator"]
                )
                separator.grid(row=row, column=0, columnspan=5, sticky="ew", pady=5)
                row += 1

        table["frame"].update_idletasks()
        table["canvas"].configure(scrollregion=table["canvas"].bbox("all"))

    def _update_stats(self, filtered_data: List[PatientData]) -> None:
        """Atualiza as estatísticas."""
        med_count = sum(1 for p in filtered_data if p.tipo == "Médico")
        psi_count = sum(1 for p in filtered_data if p.tipo == "Psicólogo")
        total = len(filtered_data)

        stats = f"Total: {total} | Médico: {med_count} | Psicólogo: {psi_count}"
        self.ui_refs["stats"].config(text=stats)

    def _filter_data(self) -> List[PatientData]:
        """Sistema de filtragem otimizado com cache de resultados."""
        filters = self.ui_refs["filters"]
        current_filters = {
            "type": filters["type"].get(),
            "search": filters["search"].get().lower(),
            "payment": filters["payment"].get().lower(),
        }

        # Verifica se os filtros mudaram
        if current_filters == self.data_cache["last_filters"]:
            return self.data_cache.get("last_result", [])

        def matches_criteria(patient: PatientData) -> bool:
            if (
                current_filters["search"]
                and current_filters["search"] not in patient.search_text
            ):
                return False
            if (
                current_filters["payment"]
                and current_filters["payment"] not in patient.pagamento.lower()
            ):
                return False
            return True

        filtered = []
        if current_filters["type"] in ["todos", "medico"]:
            filtered.extend(
                [p for p in self.data_cache["medico"] if matches_criteria(p)]
            )

        if current_filters["type"] in ["todos", "psi"]:
            filtered.extend([p for p in self.data_cache["psi"] if matches_criteria(p)])

        # Aplica ordenação
        if self.sort_state["column"] is not None:
            key_funcs = [
                lambda x: int(x.nome.split()[0]) if x.nome.split()[0].isdigit() else 0,
                lambda x: x.nome.lower(),
                lambda x: x.renach,
                lambda x: x.pagamento,
                lambda x: x.tipo,
            ]

            filtered.sort(
                key=key_funcs[self.sort_state["column"]],
                reverse=self.sort_state["reverse"],
            )

        # Atualiza cache
        self.data_cache["last_filters"] = current_filters.copy()
        self.data_cache["last_result"] = filtered

        return filtered

    def _delayed_filter(self) -> None:
        """Implementa filtragem com delay para melhor performance."""
        if self.data_cache["timer"]:
            self.master.after_cancel(self.data_cache["timer"])
        self.data_cache["timer"] = self.master.after(300, self._apply_filters)

    def _apply_filters(self) -> None:
        """Aplica filtros com animação suave."""
        if self.data_cache["timer"]:
            self.master.after_cancel(self.data_cache["timer"])

        def animate():
            filtered_data = self._filter_data()
            self._update_table(filtered_data)
            self._update_stats(filtered_data)

        self.data_cache["timer"] = self.master.after(150, animate)

    def display(self) -> None:
        """Exibe a interface principal com tratamento de erros aprimorado."""
        try:
            if not self._load_data():
                messagebox.showerror("Erro", "Não foi possível carregar os dados")
                return

            window, _ = self._create_ui()
            self._apply_filters()

            # Centralização e dimensionamento responsivo
            window.update_idletasks()
            width = max(window.winfo_width(), self.ui_config["min_width"])
            height = max(window.winfo_height(), self.ui_config["min_height"])
            x = (window.winfo_screenwidth() // 2) - (width // 2)
            y = (window.winfo_screenheight() // 2) - (height // 2)
            window.geometry(f"{width}x{height}+{x}+{y}")

            # Configuração de redimensionamento
            window.minsize(self.ui_config["min_width"], self.ui_config["min_height"])

            def on_closing():
                if self.data_cache["timer"]:
                    self.master.after_cancel(self.data_cache["timer"])
                window.destroy()

            window.protocol("WM_DELETE_WINDOW", on_closing)

        except Exception as e:
            self.logger.error(f"Erro ao exibir interface: {e}")
            messagebox.showerror(
                "Erro",
                f"Ocorreu um erro ao exibir a interface: {str(e)}\nPor favor, contate o suporte.",
            )
