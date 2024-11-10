from openpyxl import load_workbook
import PySimpleGUI as sg


class Planilhas:
    """Inicializa a classe Planilhas com um objeto de workbook."""

    def __init__(self, file_path):
        """Inicializa a classe Planilhas com um objeto de workbook."""
        try:
            self.file_path = file_path
            self.wb = load_workbook(file_path)  # Carrega a planilha
            self.sheet_name = None  # Nome da sheet ativa
            self.active_sheet = self.wb.active  # Referência à sheet ativa
        except Exception as e:
            raise ValueError(f"Erro ao carregar a planilha: {str(e)}")

    """Define e ativa uma nova sheet."""

    def set_active_sheet(self, sheet_name):
        """Define e ativa uma nova sheet."""
        try:
            if sheet_name in self.wb.sheetnames:
                self.sheet_name = sheet_name
                self.wb.active = self.wb[sheet_name]
                self.active_sheet = self.wb[sheet_name]
                self.wb.save(self.file_path)
                return True
            return False
        except Exception as e:
            print(f"Erro ao definir sheet ativa: {e}")
            return False

    """Retorna a sheet ativa atual."""

    def get_active_sheet(self):
        """Retorna a sheet ativa atual."""
        if self.sheet_name and self.sheet_name in self.wb.sheetnames:
            return self.wb[self.sheet_name]
        return self.wb.active

    """Recarrega o workbook do arquivo."""

    def reload_workbook(self):
        """Recarrega o workbook do arquivo."""
        try:
            self.wb = load_workbook(self.file_path)
            if self.sheet_name and self.sheet_name in self.wb.sheetnames:
                self.active_sheet = self.wb[self.sheet_name]
            else:
                self.active_sheet = self.wb.active
            return True
        except Exception as e:
            print(f"Erro ao recarregar workbook: {e}")
            return False

    """Conta o número de pessoas e a quantidade de pagamentos."""

    def contar_pagamento(self, col_inicial, col_final):
        """Conta o número de pessoas e a quantidade de pagamentos."""
        n_pessoa = 0
        cont_pag = {"D": 0, "C": 0, "E": 0, "P": 0}

        # Usa a sheet ativa correta
        ws = self.get_active_sheet()

        for row in ws.iter_rows(
            min_row=3, max_row=ws.max_row, min_col=col_inicial, max_col=col_final
        ):
            nome = row[0]
            if isinstance(nome.value, str) and nome.value.strip():
                n_pessoa += 1

            pag = row[4]
            if pag and pag.value in cont_pag:
                cont_pag[pag.value] += 1

        return n_pessoa, cont_pag

    """Conta a quantidade de pessoas e pagamentos feitos por médicos."""

    def contar_medico(self):
        """Conta a quantidade de pessoas e pagamentos feitos por médicos."""
        return self.contar_pagamento(2, 6)

    """Conta a quantidade de pessoas e pagamentos feitos por psicólogos."""

    def contar_psi(self):
        """Conta a quantidade de pessoas e pagamentos feitos por psicólogos."""
        return self.contar_pagamento(8, 12)

    """Exibe os resultados de contagem para médicos e psicólogos."""

    def exibir_resultado(self, janela_menu):
        """Exibe os resultados de contagem para médicos e psicólogos."""
        n_medico, pag_medico = self.contar_medico()
        n_psicologo, pag_psicologo = self.contar_psi()

        layout_resultado = (
            [
                [sg.Text("MÉDICO:")],
                [sg.Text(f"Número de pacientes: {n_medico}")],
                [sg.Text("Formas de pagamento:")],
            ]
            + [
                [
                    sg.Text(f"{tipo_pagamento}: {quantidade}")
                    for tipo_pagamento, quantidade in pag_medico.items()
                ]
            ]
            + [
                [sg.Text("")],
                [sg.Text("PSICÓLOGO:")],
                [sg.Text(f"Número de pacientes: {n_psicologo}")],
                [sg.Text("Formas de pagamento:")],
            ]
            + [
                [
                    sg.Text(f"{tipo_pagamento}: {quantidade}")
                    for tipo_pagamento, quantidade in pag_psicologo.items()
                ]
            ]
            + [[sg.Button("Voltar")]]
        )

        janela_resultado = sg.Window("Resultados Drs", layout_resultado)
        janela_menu.hide()

        while True:
            eventos, valores = janela_resultado.read()
            if eventos in (sg.WIN_CLOSED, "Voltar"):
                break

        janela_resultado.close()

        janela_menu.un_hide()

    """Coleta informações dos pacientes (médicos e psicólogos)."""

    def processar_informacao(self):
        """Coleta informações dos pacientes (médicos e psicólogos).

        Returns:
            list: Lista com informações dos pacientes.
        """
        info = []
        ws = self.wb.active

        # Processar informações de médicos e psicólogos
        for col_range in [(2, 6), (8, 12)]:
            for row in ws.iter_rows(
                min_row=3,
                max_row=ws.max_row,
                min_col=col_range[0],
                max_col=col_range[1],
            ):
                linha = [
                    cell.value
                    for cell in row
                    if isinstance(cell.value, (str, int)) and str(cell.value).strip()
                ]
                if (
                    linha
                ):  # Adiciona a linha à lista de informações se não estiver vazia
                    info.append(linha)
        return info

    """Exibe informações dos pacientes."""

    def exibir_informacao(self, janela_menu):
        """Exibe informações dos pacientes."""
        # Usa a sheet ativa correta
        ws = self.get_active_sheet()
        medico, psi = [], []

        # Informações de médicos
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=6):
            linha = [
                cell.value
                for cell in row
                if isinstance(cell.value, (str, int)) and str(cell.value).strip()
            ]
            if linha:
                medico.append(linha)

        # Informações de psicólogos
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=8, max_col=12):
            linha = [
                cell.value
                for cell in row
                if isinstance(cell.value, (str, int)) and str(cell.value).strip()
            ]
            if linha:
                psi.append(linha)

        layout_informacao = []

        # Exibe informações de médicos
        if medico:
            layout_informacao.append([sg.Text("MEDICO:")])
            for i, paciente in enumerate(medico, start=1):
                layout_informacao.append([sg.Text(f"{i} - {paciente}")])
            layout_informacao.append([sg.Text("")])

        # Exibe informações de psicólogos
        if psi:
            layout_informacao.append([sg.Text("PSICOLOGO:")])
            for i, paciente in enumerate(psi, start=1):
                layout_informacao.append([sg.Text(f"{i} - {paciente}")])

        layout_informacao.append([sg.Button("Voltar")])

        janela_informacao = sg.Window("Informação dos Pacientes", layout_informacao)
        janela_menu.hide()

        while True:
            eventos, valores = janela_informacao.read()
            if eventos in (sg.WIN_CLOSED, "Fechar"):
                break
            if eventos == "Voltar":
                janela_informacao.close()
                janela_menu.un_hide()
                break

    """Adiciona uma nova informação de paciente ao Excel com uma interface gráfica."""

    def adicionar_informacao(self, janela_menu):
        """Adiciona uma nova informação de paciente ao Excel com uma interface gráfica."""
        ws = self.get_active_sheet()

        layout = [
            [sg.Text("Deseja adicionar informações para:")],
            [
                sg.Radio("Médico", "OPCAO", key="medico"),
                sg.Radio("Psicólogo", "OPCAO", key="psicologo"),
                sg.Radio("Ambos", "OPCAO", key="ambos"),
            ],
            [sg.Text("Nome:"), sg.InputText(key="nome")],
            [sg.Text("Renach:"), sg.InputText(key="renach")],
            [sg.Text("Forma de Pagamento:")],
            [
                sg.Radio("D", "PAGAMENTO", key="debito"),
                sg.Radio("C", "PAGAMENTO", key="credito"),
                sg.Radio("E", "PAGAMENTO", key="dinheiro"),
                sg.Radio("P", "PAGAMENTO", key="pix"),
            ],
            [sg.Button("Adicionar"), sg.Button("Voltar")],
        ]

        window = sg.Window("Adicionar Informação de Paciente", layout)

        janela_menu.hide()

        while True:
            event, values = window.read()
            if event in (sg.WINDOW_CLOSED, "Voltar"):
                window.close()
                janela_menu.un_hide()
                break

            if event == "Adicionar":
                nome = values["nome"].strip().upper()
                renach = values["renach"].strip()

                # Verifica se o nome e renach estão preenchidos
                if not nome or not renach:
                    sg.popup_error("Por favor, preencha todos os campos.")
                    continue

                # Verifica se o renach é um número inteiro
                if not renach.isdigit():
                    sg.popup_error("Renach deve ser um número inteiro.")
                    continue

                # Mapeia as formas de pagamento
                formas_pagamento = {
                    "debito": "D",
                    "credito": "C",
                    "dinheiro": "E",
                    "pix": "P",
                }

                forma_pag = next(
                    (formas_pagamento[key] for key in formas_pagamento if values[key]),
                    None,
                )

                if not forma_pag:
                    sg.popup_error("Por favor, selecione uma forma de pagamento.")
                    continue

                escolha = (
                    "1" if values["medico"] else "2" if values["psicologo"] else "3"
                )

                # Encontra a próxima linha vazia em qualquer coluna relevante
                nova_linha_medico = next(
                    (
                        row
                        for row in range(3, ws.max_row + 2)
                        if not ws[f"B{row}"].value
                    ),
                    None,
                )
                nova_linha_psicologo = next(
                    (
                        row
                        for row in range(3, ws.max_row + 2)
                        if not ws[f"H{row}"].value
                    ),
                    None,
                )

                # Adiciona as informações do paciente com base na escolha
                if escolha in ["1", "3"]:
                    ws[f"B{nova_linha_medico}"] = nome
                    ws[f"C{nova_linha_medico}"] = renach
                    ws[f"F{nova_linha_medico}"] = forma_pag
                    sg.popup("Informações de médico adicionadas com sucesso!")

                if escolha in ["2", "3"]:
                    ws[f"H{nova_linha_psicologo}"] = nome
                    ws[f"I{nova_linha_psicologo}"] = renach
                    ws[f"L{nova_linha_psicologo}"] = forma_pag
                    sg.popup("Informações de psicólogo adicionadas com sucesso!")

                self.wb.save("CAMPSSA.xlsx")  # Salva as alterações

    """Remove informações de pacientes da planilha com base no RENACH fornecido pelo usuário."""

    def excluir(self, janela_menu):
        """Remove informações de pacientes da planilha com base no RENACH fornecido pelo usuário."""
        ws = self.get_active_sheet()
        pacientes_medicos = {}
        pacientes_psicologos = {}

        # Armazenar pacientes de médicos e psicólogos
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            # Pacientes médicos
            if row[1].value and row[2].value:
                try:
                    renach_medico = int(row[2].value)
                    pacientes_medicos.setdefault(renach_medico, []).append(row[0].row)
                except ValueError:
                    print(
                        f"RENACH inválido na linha {row[0].row}: {row[2].value}"
                    )

            # Pacientes psicólogos
            if row[7].value and row[8].value:
                try:
                    renach_psicologo = int(row[8].value)
                    pacientes_psicologos.setdefault(renach_psicologo, []).append(
                        row[0].row
                    )
                except ValueError:
                    print(
                        f"RENACH inválido na linha {row[0].row}: {row[8].value}"
                    )

        layout = [
            [sg.Text("Informe o RENACH do paciente para exclusão:")],
            [sg.InputText(key="renach")],
            [sg.Button("Excluir"), sg.Button("Voltar")],
        ]

        window = sg.Window("Excluir Paciente", layout)
        janela_menu.hide()

        while True:
            event, values = window.read()
            if event in (sg.WINDOW_CLOSED, "Voltar"):
                window.close()
                janela_menu.un_hide()
                break

            if event == "Excluir":
                renach_input = values["renach"].strip()

                # Validar entrada do RENACH
                try:
                    renach = int(renach_input)
                except ValueError:
                    sg.popup_error("RENACH deve ser um número inteiro.")
                    continue

                paciente_removido = (
                    False  # Para rastrear se algum paciente foi removido
                )

                # Limpar informações de pacientes médicos se o RENACH existir
                if renach in pacientes_medicos:
                    for linha_a_limpar in pacientes_medicos[renach]:
                        # Limpa o nome do médico
                        ws[f"B{linha_a_limpar}"] = None
                        # Limpa o RENACH do médico
                        ws[f"C{linha_a_limpar}"] = None
                        # Limpa a forma de pagamento do médico
                        ws[f"F{linha_a_limpar}"] = None
                        paciente_removido = True

                # Limpar informações de pacientes psicólogos se o RENACH existir
                if renach in pacientes_psicologos:
                    for linha_a_limpar in pacientes_psicologos[renach]:
                        # Limpa o nome do psicólogo
                        ws[f"H{linha_a_limpar}"] = None
                        # Limpa o RENACH do psicólogo
                        ws[f"I{linha_a_limpar}"] = None
                        # Limpa a forma de pagamento do psicólogo
                        ws[f"L{linha_a_limpar}"] = None
                        paciente_removido = True

                # Salva as alterações se houve limpezas
                if paciente_removido:
                    self.wb.save("CAMPSSA.xlsx")
                    sg.popup("Informações de pacientes removidas com sucesso!")
                else:
                    sg.popup_error("RENACH inválido ou paciente não encontrado.")

    def valores_totais(self, janela_menu):
        self.reload_workbook()

        n_medico, pag_medico = self.contar_medico()
        n_psicologo, pag_psicologo = self.contar_psi()

        total_medico = n_medico * 148.65
        total_psicologo = n_psicologo * 192.61

        valor_medico = n_medico * 49
        valor_psicologo = n_psicologo * 63.50

        layout = [
            [sg.Text("MEDICO")],
            [sg.Text(f"Valor total: {total_medico:.2f}")],
            [sg.Text(f"Valor a ser pago: {valor_medico:.2f}")],
            [sg.Text("")],
            [sg.Text("PSICOLOGO")],
            [sg.Text(f"Valor total: {total_psicologo:.2f}")],
            [sg.Text(f"Valor a ser pago: {valor_psicologo:.2f}")],
            [sg.Button(f"Voltar")],
        ]

        janela = sg.Window("Contas", layout)

        janela_menu.hide()

        while True:
            eventos, valores = janela.read()
            if eventos in (sg.WINDOW_CLOSED, "Voltar"):
                janela.close()
                janela_menu.un_hide()
                break
